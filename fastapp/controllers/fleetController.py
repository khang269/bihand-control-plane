from fastapi import APIRouter, Depends, HTTPException, BackgroundTasks, File, UploadFile
from fastapi.responses import RedirectResponse
from pydantic import BaseModel, Field
from typing import List, Optional, Dict, Any
import logging
import os
import secrets
import csv
import io
from datetime import datetime, timezone
from urllib.parse import urlencode
from openpyxl import load_workbook

import requests

from fastapp.controllers.authController import get_current_user
from fastapp.models.fleetModel import FleetModel
from fastapp.models.instanceModel import InstanceModel
from fastapp.models.chatMessageModel import ChatMessageModel
from fastapp.services.agentProfileService import (
    DEFAULT_AGENT_MD,
    adapter_capabilities,
    build_skill_snapshot,
    sync_skills,
)

# Log configuration
logger = logging.getLogger(__name__)

fleetRouter = APIRouter(tags=["Fleet / Bihand Company"])

@fleetRouter.post("/parse-roster", summary="Parse roster CSV/XLSX and map to org json hierarchy")
async def parse_roster_file(file: UploadFile = File(...), auth_payload: dict = Depends(get_current_user)):
    file_bytes = await file.read()
    raw_agents = []
    global_custom_skills = {}
    
    try:
        # Parse Excel (XLSX)
        if file.filename.endswith('.xlsx') or file.filename.endswith('.xls'):
            wb = load_workbook(io.BytesIO(file_bytes), data_only=True)
            
            # 1. Parse custom skills sheet if it exists
            skills_sheet = None
            for sheetname in wb.sheetnames:
                if sheetname.strip().lower() in ["skills", "customskills", "custom_skills", "skills_files"]:
                    skills_sheet = wb[sheetname]
                    break
                    
            if skills_sheet:
                skills_rows = list(skills_sheet.iter_rows(values_only=True))
                if skills_rows and len(skills_rows) > 1:
                    headers = [str(h).strip().lower() for h in skills_rows[0] if h is not None]
                    name_idx, content_idx = -1, -1
                    for idx, h in enumerate(headers):
                        if h in ["name", "skill", "skillname", "key"]:
                            name_idx = idx
                        elif h in ["content", "instructions", "description", "body"]:
                            content_idx = idx
                            
                    if name_idx != -1 and content_idx != -1:
                        for row in skills_rows[1:]:
                            if row and len(row) > max(name_idx, content_idx):
                                sk_name = str(row[name_idx]).strip() if row[name_idx] is not None else ""
                                sk_content = str(row[content_idx]).strip() if row[content_idx] is not None else ""
                                if sk_name:
                                    global_custom_skills[sk_name.strip().lower()] = sk_content
            
            # 2. Parse main agents sheet
            sheet = wb.active
            rows = list(sheet.iter_rows(values_only=True))
            if not rows:
                raise HTTPException(status_code=400, detail="Uploaded Excel file is empty.")
            
            headers = [str(h).strip().lower() for h in rows[0] if h is not None]
            for row in rows[1:]:
                if not any(row):  # Skip empty rows
                    continue
                row_dict = {}
                for idx, val in enumerate(row):
                    if idx < len(headers):
                        row_dict[headers[idx]] = val
                raw_agents.append(row_dict)
                
        # Parse CSV
        else:
            text_content = file_bytes.decode('utf-8', errors='ignore')
            
            # Check for multi-section format (e.g. separated by [Skills], --- SKILLS ---, or similar)
            import re
            parts = re.split(r'\n\s*(?:\[skills\]|---+\s*skills\s*---+|===+\s*skills\s*===+)\s*\n', text_content, flags=re.IGNORECASE)
            
            agents_csv = parts[0]
            skills_section = parts[1] if len(parts) > 1 else None
            
            reader = csv.DictReader(io.StringIO(agents_csv))
            for row in reader:
                if not any(row.values()):
                    continue
                row_dict = {k.strip().lower(): v for k, v in row.items() if k is not None}
                raw_agents.append(row_dict)
                
            if skills_section:
                # Parse as custom markdown sections (highly user-friendly: no quotes or CSV column constraints)
                current_skill_name = None
                current_skill_lines = []
                
                for line in skills_section.splitlines():
                    stripped_line = line.strip()
                    # Match headers like "# web-audit", "## web-audit", or "[web-audit]"
                    header_match = re.match(r'^(?:#+|\[+)\s*([a-zA-Z0-9_\-]+)\s*(?:\]+)?$', stripped_line)
                    if header_match:
                        # Save previous skill
                        if current_skill_name and current_skill_lines:
                            global_custom_skills[current_skill_name.strip().lower()] = "\n".join(current_skill_lines).strip()
                        current_skill_name = header_match.group(1).strip()
                        current_skill_lines = []
                    else:
                        if current_skill_name is not None:
                            # Keep raw line contents to preserve user's markdown lists, layout, and spacing
                            current_skill_lines.append(line)
                            
                # Save the last skill
                if current_skill_name and current_skill_lines:
                    global_custom_skills[current_skill_name.strip().lower()] = "\n".join(current_skill_lines).strip()
                
    except Exception as e:
        logger.error(f"Error parsing file: {e}")
        raise HTTPException(status_code=400, detail=f"Failed to parse uploaded file: {str(e)}")
        
    if not raw_agents:
        raise HTTPException(status_code=400, detail="No agent data found in file.")

    # Normalize and resolve hierarchy
    role_to_id = {}
    normalized_agents = []
    default_models = {
        'anthropic': 'claude-sonnet-4-6',
        'openai': 'gpt-5.5',
        'gemini': 'gemini-2.5-flash',
        'deepseek': 'deepseek-chat',
    }

    # Step 1: Extract global custom skill definitions from standard rows or columns
    system_skills_set = {
        'bihand-agent', 'bihand-browser-use', 'bihand', 'bihand-dev', 'bihand-create-agent',
        'bihand-google-workspace', 'meta-mcp', 'social-instagram',
        'social-x', 'social-reddit'
    }

    # Separate raw agent rows from dedicated skill definition rows
    filtered_agent_rows = []
    for raw in raw_agents:
        role_val = str(raw.get('role', '')).strip().lower()
        # If the row has role='skill' or role='custom-skill', it is a global skill definition row!
        if role_val in ['skill', 'custom-skill', 'custom_skill', 'skills_file']:
            sk_name = str(raw.get('title', raw.get('name', raw.get('role_key', '')))).strip()
            # Try fetching skill content from content or standard columns
            sk_content = str(raw.get('customskills', raw.get('custom_skills', raw.get('skills_files', raw.get('skillsfiles', raw.get('content', raw.get('instructions', raw.get('description', '')))))))).strip()
            if sk_name and sk_content:
                global_custom_skills[sk_name.strip().lower()] = sk_content.replace(r"\n", "\n")
        else:
            filtered_agent_rows.append(raw)

    for raw in filtered_agent_rows:
        # Check standard columns for definitions
        custom_skills_raw = str(raw.get('customskills', raw.get('custom_skills', raw.get('skills_files', raw.get('skillsfiles', raw.get('custom_skills_definitions', '')))))).strip()
        if custom_skills_raw:
            is_json = False
            if (custom_skills_raw.startswith('[') and custom_skills_raw.endswith(']')) or (custom_skills_raw.startswith('{') and custom_skills_raw.endswith('}')):
                try:
                    import json
                    parsed_json = json.loads(custom_skills_raw)
                    is_json = True
                    if isinstance(parsed_json, list):
                        for item in parsed_json:
                            if isinstance(item, dict) and "name" in item:
                                global_custom_skills[str(item["name"]).strip().lower()] = str(item.get("content", "")).strip()
                    elif isinstance(parsed_json, dict):
                        for k, v in parsed_json.items():
                            global_custom_skills[k.strip().lower()] = str(v).strip()
                except Exception:
                    pass
            
            if not is_json:
                parts = custom_skills_raw.split("||")
                for part in parts:
                    if "::" in part:
                        sk_name, sk_content = part.split("::", 1)
                        sk_name = sk_name.strip()
                        sk_content = sk_content.strip().replace(r"\n", "\n")
                        if sk_name:
                            global_custom_skills[sk_name.strip().lower()] = sk_content
        
        # Check for column keys matching pattern "skill:something"
        for k, v in raw.items():
            if k and str(k).startswith("skill:") and v:
                sk_name = str(k).split(":", 1)[1].strip()
                sk_content = str(v).strip().replace(r"\n", "\n")
                if sk_name:
                    global_custom_skills[sk_name.strip().lower()] = sk_content

    # Step 2: Validate duplicate role names in filtered agent rows
    seen_roles = {}
    for idx, raw in enumerate(filtered_agent_rows):
        role = str(raw.get('role', '')).strip()
        if not role:
            role = f"Employee {idx + 1}"
        role_upper = role.upper()
        if role_upper in seen_roles:
            seen_roles[role_upper] += 1
            new_role = f"{role} {seen_roles[role_upper]}"
            raw['role'] = new_role
        else:
            seen_roles[role_upper] = 1

    # Step 3: Generate UUIDs, resolve skills to system vs custom, and register roles
    for idx, raw in enumerate(filtered_agent_rows):
        role = str(raw.get('role', '')).strip()
        if not role:
            role = f"Employee {idx + 1}"
            
        agent_id = f"agent_{role.lower().replace(' ', '_')}_{secrets.token_hex(4)}"
        role_to_id[role.upper()] = agent_id
        
        title = str(raw.get('title', '')).strip() or f"{role} Staff"
        
        agent_type = str(raw.get('agenttype', '')).strip().lower() or 'opencode'
        if agent_type not in ['opencode', 'openclaw', 'claudecode']:
            agent_type = 'opencode'
            
        provider = str(raw.get('provider', '')).strip().lower() or 'anthropic'
        if provider == 'google':
            provider = 'gemini'
        if provider not in ['anthropic', 'openai', 'gemini', 'deepseek']:
            provider = 'anthropic'
            
        model = str(raw.get('model', '')).strip()
        if not model:
            model = default_models.get(provider, 'claude-sonnet-4-6')
            
        machine_type = str(raw.get('machinetype', '')).strip().lower() or 'e2-small'
        if machine_type not in ['e2-small', 'e2-medium', 'e2-standard-2']:
            machine_type = 'e2-small'
            
        avatar_hash = str(raw.get('avatarhash', '')).strip()
        if not avatar_hash:
            avatar_hash = None
                
        # Parse comma-separated list of skill names/slugs
        skills_raw = str(raw.get('skills', raw.get('enabledskills', raw.get('enabled_skills', '')))).strip()
        listed_skills = [s.strip() for s in skills_raw.split(',') if s.strip()] if skills_raw else ['bihand']

        agent_enabled_system_skills = []
        agent_custom_skills_dict = {}

        # Parse inline custom skill definitions on this row specifically, if any
        row_custom_skills_raw = str(raw.get('customskills', raw.get('custom_skills', raw.get('skills_files', raw.get('skillsfiles', ''))))).strip()
        if row_custom_skills_raw:
            is_json = False
            if (row_custom_skills_raw.startswith('[') and row_custom_skills_raw.endswith(']')) or (row_custom_skills_raw.startswith('{') and row_custom_skills_raw.endswith('}')):
                try:
                    import json
                    parsed_json = json.loads(row_custom_skills_raw)
                    is_json = True
                    if isinstance(parsed_json, list):
                        for item in parsed_json:
                            if isinstance(item, dict) and "name" in item:
                                agent_custom_skills_dict[str(item["name"]).strip()] = str(item.get("content", "")).strip()
                    elif isinstance(parsed_json, dict):
                        for k, v in parsed_json.items():
                            agent_custom_skills_dict[str(k).strip()] = str(v).strip()
                except Exception:
                    pass
            if not is_json:
                parts = row_custom_skills_raw.split("||")
                for part in parts:
                    if "::" in part:
                        sk_name, sk_content = part.split("::", 1)
                        sk_name = sk_name.strip()
                        sk_content = sk_content.strip().replace(r"\n", "\n")
                        if sk_name:
                            agent_custom_skills_dict[sk_name] = sk_content

        # Match each listed skill name to either a pre-installed system skill or custom definition
        for s in listed_skills:
            # Enforce folder-valid/filename-valid skill names (alphanumeric, dashes, underscores)
            import re
            sanitized_skill_name = re.sub(r'[^a-zA-Z0-9_\-]', '_', s).strip().lower()
            if not sanitized_skill_name:
                continue

            if sanitized_skill_name in system_skills_set:
                agent_enabled_system_skills.append(sanitized_skill_name)
            else:
                # Look up the custom skill directly using case-insensitive raw keys to preserve content
                p_names = [
                    s.strip().lower(),
                    sanitized_skill_name
                ]
                matched_content = None
                for name_var in p_names:
                    if name_var in global_custom_skills:
                        matched_content = global_custom_skills[name_var]
                        break
                        
                if matched_content is not None:
                    agent_custom_skills_dict[sanitized_skill_name] = matched_content
                else:
                    # Check if this row specifically defines the skill in a column skill:name
                    row_skill_content = None
                    for name_var in p_names:
                        row_skill_content = raw.get(f"skill:{name_var}") or raw.get(f"skill:{name_var.upper()}")
                        if row_skill_content:
                            break
                    if not row_skill_content:
                        row_skill_content = raw.get(f"skill:{s}")
                        
                    if row_skill_content and str(row_skill_content).strip():
                        agent_custom_skills_dict[sanitized_skill_name] = str(row_skill_content).strip().replace(r"\n", "\n")
                    else:
                        # If they specified a skill name but didn't define its content, fallback to a placeholder
                        if sanitized_skill_name not in agent_custom_skills_dict:
                            agent_custom_skills_dict[sanitized_skill_name] = f"# {sanitized_skill_name}\nCustom instructions for {sanitized_skill_name}."

        # Guarantee bihand is included
        if "bihand" not in agent_enabled_system_skills:
            agent_enabled_system_skills.insert(0, "bihand")

        custom_skills_list = [{"name": name, "content": content} for name, content in agent_custom_skills_dict.items()]
        
        # Parse customAgentMd instruction override column from roster CSV/XLSX
        user_custom_instructions = str(raw.get('agentmd', raw.get('agent_md', raw.get('customagentmd', raw.get('custom_agent_md', raw.get('instructions', '')))))).strip()
        
        normalized_agents.append({
            "id": agent_id,
            "role": role,
            "title": title,
            "agentType": agent_type,
            "provider": provider,
            "apiKey": "",  # UI links credential mapping
            "model": model,
            "machineType": machine_type,
            "reportsTo": str(raw.get('reportsto', '')).strip(),
            "enabledSkills": agent_enabled_system_skills,
            "avatarHash": avatar_hash,
            "skillsFiles": custom_skills_list,
            "customAgentMd": user_custom_instructions
        })

    # Step 4: Establish parent manager ID references and check/resolve cycle dependencies
    for ag in normalized_agents:
        raw_reports_to = ag["reportsTo"]
        if raw_reports_to:
            ag["reportsTo"] = role_to_id.get(raw_reports_to.upper(), None)
        else:
            ag["reportsTo"] = None

    # Detect and resolve circular reporting structures (A -> B -> C -> A)
    for ag in normalized_agents:
        if not ag["reportsTo"]:
            continue
        visited = set()
        curr = ag
        has_cycle = False
        while curr["reportsTo"]:
            if curr["id"] in visited:
                has_cycle = True
                break
            visited.add(curr["id"])
            parent = next((parent_agent for parent_agent in normalized_agents if parent_agent["id"] == curr["reportsTo"]), None)
            if not parent:
                break
            curr = parent
        if has_cycle:
            # Break cycle immediately by resetting direct reference back to Board / None
            ag["reportsTo"] = None

    return {"success": True, "agents": normalized_agents}

    return {"success": True, "agents": normalized_agents}


def _normalized_agent_md(value: Optional[str]) -> str:
    content = (value or "").strip()
    if not content:
        return DEFAULT_AGENT_MD
    if content.startswith("You are a Bihand autonomous corporate agent."):
        return DEFAULT_AGENT_MD
    if content.startswith("You are an autonomous worker agent in a fast-moving AI startup."):
        return DEFAULT_AGENT_MD
    if "No such file or directory" in content or content.startswith("cat: "):
        return DEFAULT_AGENT_MD
    return content

class AgentConfig(BaseModel):
    id: Optional[str] = Field(default=None, description="Temporary ID for graph relationships")
    role: str = Field(..., description="Role of the agent (e.g. CEO, CTO)")
    title: Optional[str] = Field(default=None, description="Job title")
    reportsTo: Optional[str] = Field(default=None, description="Temporary ID of manager agent")
    agentType: str = Field(..., description="Type of agent (e.g. openclaw, opencode, claudecode, hermes)")
    provider: str = Field(..., description="LLM provider (e.g. anthropic, openai)")
    apiKey: Optional[str] = Field(default="", description="API key credential ID for the provider. Not required when oauthToken is set.")
    oauthToken: Optional[str] = Field(default=None, description="Subscription auth instead of apiKey: for claudecode, a `claude setup-token` OAuth token; for codex, the full pasted contents of a `~/.codex/auth.json` file from `codex login`.")
    customBaseUrl: Optional[str] = Field(default=None, description="Custom provider only (provider='custom'): base URL of an arbitrary OpenAI-compatible endpoint to point the agent at instead of a named provider.")
    model: Optional[str] = Field(default=None)
    machineType: str = Field(default="e2-small", description="GCP Machine Type")
    agentMd: Optional[str] = Field(default="", description="Markdown persona")
    customAgentMd: Optional[str] = Field(default="", description="User custom instruction override")
    soulMd: Optional[str] = Field(default="", description="Agent soul prompt")
    toolsMd: Optional[str] = Field(default="", description="Tooling instructions")
    mcpConfig: Optional[str] = Field(default="", description="MCP server config")
    enabledSkills: Optional[List[str]] = Field(default=[], description="Enabled skill slugs")
    skillsFiles: Optional[List[Dict[str, str]]] = Field(default=[], description="Custom skill files with name and content")
    durationMonths: int = Field(default=1, description="Duration in months for this agent")
    durationDays: Optional[int] = Field(default=30, description="Duration in days for fine-grained tryout packages")
    avatarHash: Optional[str] = Field(default=None, description="The 3D avatar model hash")

class UpdateBudgetRequest(BaseModel):
    apiBudget: float = Field(..., description="New Monthly API Spend limit (USD)")

@fleetRouter.patch("/{fleet_id}/budget", summary="Update Fleet API Budget")
async def update_fleet_budget(fleet_id: str, req: UpdateBudgetRequest, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    
    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")
        
    FleetModel._collection().update_one({"_id": fleet_id}, {"$set": {"apiBudget": req.apiBudget}})
    return {"message": "Budget updated successfully", "apiBudget": req.apiBudget}

@fleetRouter.delete("/{fleet_id}", summary="Destroy entire fleet and resources")
async def delete_fleet(fleet_id: str, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    
    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")
        
    if fleet.get("status") in ["deleting", "deleted"]:
        raise HTTPException(status_code=400, detail="Fleet is already being destroyed")
        
    # Mark fleet as deleting
    FleetModel._updateStatus(fleet_id, "deleting")
    
    from fastapp.database import get_db
    # Instantly update statuses and broadcast changes via WebSockets so the UI transitions immediately
    instances_cursor = get_db()["instances"].find({"fleetId": fleet_id, "status": {"$nin": ["deleted"]}})
    for inst in instances_cursor:
        try:
            InstanceModel._updateStatus(str(inst["_id"]), "deleting")
        except Exception as ws_err:
            logger.warning(f"Failed to broadcast individual deletion status for {inst.get('_id')}: {ws_err}")
            # Fallback to direct DB write if WS fails
            get_db()["instances"].update_one({"_id": inst["_id"]}, {"$set": {"status": "deleting"}})
    
    from fastapp.tasks import delete_fleet_task
    delete_fleet_task.delay(fleet_id)
        
    return {"message": "Fleet destruction initiated"}

class AgentConfigPayload(BaseModel):
    agentMd: str = Field(..., description="Agent System Prompt / Identity Markdown")
    soulMd: Optional[str] = Field(default="", description="Agent soul markdown")
    toolsMd: Optional[str] = Field(default="", description="Tool instructions markdown")
    mcpConfig: str = Field(..., description="MCP Servers JSON Configuration")
    enabledSkills: Optional[List[str]] = Field(default=[], description="Enabled skill slugs")


class AgentSkillSyncPayload(BaseModel):
    desiredSkills: List[str] = Field(default_factory=list, description="Desired skill references")


class UpdateInstructionsBundlePayload(BaseModel):
    mode: Optional[str] = Field(default=None, description="managed or external")
    rootPath: Optional[str] = Field(default=None, description="External bundle root path")
    entryFile: Optional[str] = Field(default=None, description="Entry markdown file")


class UpsertInstructionsFilePayload(BaseModel):
    path: str = Field(..., description="Relative file path inside bundle")
    content: str = Field(..., description="File content")


GOOGLE_TOOL_KEY = "googleWorkspace"
META_MCP_TOOL_KEY = "meta_mcp"
META_MCP_SERVER_NAME = "meta"
META_MCP_SKILL_SLUG = "meta-mcp"
META_DEVTOOLS_TOOL_KEY = "meta_devtools"
META_DEVTOOLS_SERVER_NAME = "meta-devtools"
# MCP server names injected systematically by the platform (chrome-devtools at provision time,
# meta via the dedicated Meta MCP integration card, meta-devtools via the Meta Developer Tools
# MCP integration card) - never editable/deletable through the generic per-agent MCP server
# management endpoints below.
PROTECTED_MCP_SERVER_NAMES = {"chrome-devtools", META_MCP_SERVER_NAME, META_DEVTOOLS_SERVER_NAME}
DEFAULT_GOOGLE_SCOPES = [
    "https://www.googleapis.com/auth/drive.readonly",
    "https://www.googleapis.com/auth/gmail.readonly",
    "https://www.googleapis.com/auth/calendar.readonly",
    "openid",
    "email",
    "profile",
]


def _google_redirect_uri() -> str:
    configured = os.environ.get("GOOGLE_WORKSPACE_REDIRECT_URI")
    if configured:
        return configured
    base_url = os.environ.get("BIHAND_PUBLIC_API_URL", "http://localhost:8000").rstrip("/")
    return f"{base_url}/api/fleets/integrations/google/callback"


def _public_web_url() -> str:
    return os.environ.get("BIHAND_WEB_URL", "http://localhost:3100").rstrip("/")


def _sanitized_tool_connections(instance: dict) -> Dict:
    source = instance.get("toolConnections") or {}
    sanitized = {}
    for key, conn in source.items():
        if not isinstance(conn, dict):
            continue
        sanitized[key] = {
            "status": conn.get("status", "not_connected"),
            "connectedAt": conn.get("connectedAt"),
            "email": conn.get("email"),
            "scopes": conn.get("scopes", []),
            "lastError": conn.get("lastError"),
            "credentialId": conn.get("credentialId"),
            "name": conn.get("name"),
        }
    return sanitized

@fleetRouter.get("/{fleet_id}/instances/{instance_id}/config", summary="Get Agent Configuration")
async def get_agent_config(fleet_id: str, instance_id: str, live: bool = False, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")
        
    instance = InstanceModel._getByIdWithKeys(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")
        
    ip = instance.get("externalIp")
    if not live or not ip or instance.get("status") != "running":
        # Return instantly from DB (highly optimized)
        skill_snapshot = build_skill_snapshot(instance)
        return {
            "agentMd": _normalized_agent_md(instance.get("agentMd", "")),
            "soulMd": instance.get("soulMd", ""),
            "toolsMd": instance.get("toolsMd", ""),
            "mcpConfig": instance.get("mcpConfig", ""),
            "enabledSkills": instance.get("enabledSkills", []),
            "desiredSkills": skill_snapshot.get("desiredSkills", []),
            "adapterCapabilities": adapter_capabilities(instance),
            "toolConnections": _sanitized_tool_connections(instance),
        }
        
    from fastapp.services import sshService
    from fastapp.services.provisionerService import get_provisioning_strategy
    private_key = instance["sshKeyPrivate"]
    agent_type = instance.get("iteration", "openclaw")
    strategy = get_provisioning_strategy(agent_type)
    
    try:
        vm_instructions = strategy.getInstructions(ip, private_key)
        agent_md = ""
        target_file_name = strategy.get_instructions_matrix().get("agentMd", "AGENTS.md")
        for f in vm_instructions:
            if f.get("name") == target_file_name:
                agent_md = f.get("content", "").strip()
                break
                
        mcp_config = strategy.getMcpConfig(ip, private_key)

        # If remote is completely empty, fallback to DB
        if not agent_md: agent_md = instance.get("agentMd", "")
        if not mcp_config: mcp_config = instance.get("mcpConfig", "")

        # A live SSH read returns the VM's real config with any bound credential's
        # secret already resolved in place - mask it back to ${cred:<key>} form before
        # this ever gets cached to Mongo or returned to the frontend.
        from fastapp.utils.mcpCredentials import mask_mcp_config_secrets
        mcp_config = mask_mcp_config_secrets(instance, mcp_config)

        # Save back to MongoDB to ensure easy recovery and cache
        try:
            from bson import ObjectId
            from fastapp.database import get_db
            get_db()["instances"].update_one(
                {"_id": ObjectId(instance_id)},
                {"$set": {"agentMd": agent_md, "mcpConfig": mcp_config, "updatedDate": datetime.now(timezone.utc)}}
            )
        except Exception as db_e:
            logger.error(f"Failed to backup VM config to MongoDB: {db_e}")

        skill_snapshot = build_skill_snapshot(instance)
        
        return {
            "agentMd": _normalized_agent_md(agent_md),
            "soulMd": instance.get("soulMd", ""),
            "toolsMd": instance.get("toolsMd", ""),
            "mcpConfig": mcp_config,
            "enabledSkills": instance.get("enabledSkills", []),
            "desiredSkills": skill_snapshot.get("desiredSkills", []),
            "adapterCapabilities": adapter_capabilities(instance),
            "toolConnections": _sanitized_tool_connections(instance),
        }
    except Exception as e:
        # Fallback to DB on SSH error
        skill_snapshot = build_skill_snapshot(instance)
        return {
            "agentMd": _normalized_agent_md(instance.get("agentMd", "")),
            "soulMd": instance.get("soulMd", ""),
            "toolsMd": instance.get("toolsMd", ""),
            "mcpConfig": instance.get("mcpConfig", ""),
            "enabledSkills": instance.get("enabledSkills", []),
            "desiredSkills": skill_snapshot.get("desiredSkills", []),
            "adapterCapabilities": adapter_capabilities(instance),
            "toolConnections": _sanitized_tool_connections(instance),
        }

@fleetRouter.get("/{fleet_id}/instances/{instance_id}/logs", summary="Get Instance Startup Logs from GCP Serial")
async def get_instance_logs(fleet_id: str, instance_id: str, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    
    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")
        
    instance = InstanceModel._getByIdWithKeys(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")
        
    from fastapp.services import gcpService
    
    logs = instance.get("startupLogs", "")
    
    # Try fetching fresh logs from GCP if VM exists
    try:
        if instance.get("status") not in ["deleted"]:
            # Our provisioning strategies execute: exec > >(tee -a /var/log/startup.log /dev/ttyS1) 2>&1
            # Therefore, the script output goes exclusively to Serial Port 2. Port 1 only has OS/Boot logs.
            raw_logs = gcpService.get_instance_serial_port_output(instance["vmName"], instance["zone"], port=2)
            
            # Find the actual bash script execution in the massive serial output
            if "=== Bihand Fleet Worker Startup ===" in raw_logs:
                logs = raw_logs[raw_logs.find("=== Bihand Fleet Worker Startup ==="):]
            elif "=== OpenCode Worker Startup ===" in raw_logs:
                logs = raw_logs[raw_logs.find("=== OpenCode Worker Startup ==="):]
            elif "=== Claude Code Worker Startup ===" in raw_logs:
                logs = raw_logs[raw_logs.find("=== Claude Code Worker Startup ==="):]
            elif "=== minerClaw Autonomous Startup" in raw_logs:
                logs = raw_logs[raw_logs.find("=== minerClaw Autonomous Startup"):]
            else:
                # If port 2 is totally empty, fallback to port 1 just in case
                if not raw_logs.strip():
                    raw_logs = gcpService.get_instance_serial_port_output(instance["vmName"], instance["zone"], port=1)
                logs = raw_logs[-5000:] # Just return the last 5000 chars if headers missing
                
            # Cache it in DB
            InstanceModel._updateStatus(instance_id, instance["status"], startupLogs=logs)
    except Exception as e:
        logger.warning(f"Failed to fetch live logs for {instance['vmName']}: {e}")
        pass
        
    return {"logs": logs, "provisionLog": instance.get("provisionLog", [])}

@fleetRouter.put("/{fleet_id}/instances/{instance_id}/config", summary="Update Agent Configuration")
async def update_agent_config(fleet_id: str, instance_id: str, req: AgentConfigPayload, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    
    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")
        
    instance = InstanceModel._getByIdWithKeys(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")
        
    if instance.get("status") not in ["running", "stopped", "error"]:
        raise HTTPException(status_code=400, detail="Forbidden: Agent must be in 'running', 'stopped', or 'error' status to perform configuration updates.")

    if instance.get("status") == "running":
        InstanceModel._updateStatus(instance_id, "updating")

    # 1. Update the Database
    InstanceModel._updateConfig(
        instance_id,
        req.agentMd,
        req.soulMd,
        req.toolsMd,
        req.mcpConfig,
    )

    next_adapter_config, _, _ = sync_skills(instance, req.enabledSkills or [])
    InstanceModel._setAdapterConfig(instance_id, next_adapter_config)
    InstanceModel._setEnabledSkills(instance_id, req.enabledSkills or [])
    
    # 2. Push the config to the VM via SSH asynchronously
    from fastapp.tasks import push_agent_config_task
    push_agent_config_task.delay(
        instance_id,
        req.agentMd,
        req.soulMd or "",
        req.toolsMd or "",
        req.mcpConfig,
        req.enabledSkills or [],
    )
    
    return {"message": "Config update initiated"}


@fleetRouter.get("/{fleet_id}/instances/{instance_id}/adapter-capabilities", summary="Get adapter capabilities")
async def get_adapter_capabilities(fleet_id: str, instance_id: str, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")

    instance = InstanceModel._getById(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")

    return {
        "adapterType": instance.get("iteration", "openclaw"),
        "capabilities": adapter_capabilities(instance),
    }


@fleetRouter.get("/{fleet_id}/instances/{instance_id}/chat/history", summary="Get persisted interactive chat history")
async def get_agent_chat_history(fleet_id: str, instance_id: str, limit: int = 200, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"

    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")

    instance = InstanceModel._getById(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")

    messages = ChatMessageModel._listByInstance(instance_id, limit=min(max(limit, 1), 500))
    return {"messages": messages}


@fleetRouter.get("/{fleet_id}/instances/{instance_id}/skills", summary="Get agent skills snapshot")
async def get_agent_skills(fleet_id: str, instance_id: str, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    
    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")

    instance = InstanceModel._getById(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")

    ip = instance.get("externalIp")
    if instance.get("status") == "running" and ip:
        from fastapp.services.provisioning import get_provisioning_strategy
        private_key = instance["sshKeyPrivate"]
        agent_type = instance.get("iteration", "openclaw")
        strategy = get_provisioning_strategy(agent_type)
        try:
            vm_skills = strategy.getSkills(ip, private_key)
            if vm_skills and any(s.get("content", "").strip() for s in vm_skills):
                # Safe merge: do not lose existing custom skills in DB if they are missing from VM,
                # and do not resurrect deleted skills that are present on the VM but missing in the DB.
                from fastapp.services.agentProfileService import SYSTEM_SKILLS_SET
                db_skills = instance.get("skillsFiles") or []
                db_skills_dict = {s["name"]: s["content"] for s in db_skills if s.get("name")}
                
                # Overwrite/update with VM skills ONLY if they exist in DB or are system-managed
                for s in vm_skills:
                    s_name = s.get("name")
                    if s_name and (s_name in db_skills_dict or s_name in SYSTEM_SKILLS_SET):
                        db_skills_dict[s_name] = s.get("content", "")
                        
                final_skills = [{"name": name, "content": content} for name, content in db_skills_dict.items()]
                
                from bson import ObjectId
                from fastapp.database import get_db
                get_db()["instances"].update_one(
                    {"_id": ObjectId(instance_id)},
                    {"$set": {"skillsFiles": final_skills, "updatedDate": datetime.now(timezone.utc)}}
                )
                return {"files": final_skills}
            else:
                raise Exception("Fetched skills are empty or VM directory not initialized yet")
        except Exception as e:
            logger.warning(f"Failed to fetch live skills for {instance_id}: {e}")
            pass

    # Fallback: Retrieve skills from MongoDB backup
    db_skills = instance.get("skillsFiles")
    if db_skills:
        return {"files": db_skills}

    # Second Fallback: Generate the default system skills snapshot from templates
    from fastapp.services.agentProfileService import build_skill_snapshot
    skill_snapshot = build_skill_snapshot(instance)
    fallback_files = [{"name": s["runtimeName"], "content": s["content"]} for s in skill_snapshot.get("files", [])]
    return {"files": fallback_files}

class SkillFilePayload(BaseModel):
    name: str
    content: str

class UpsertSkillsPayload(BaseModel):
    files: List[SkillFilePayload]

@fleetRouter.put("/{fleet_id}/instances/{instance_id}/skills", summary="Write skills")
async def put_agent_skills_list(
    fleet_id: str,
    instance_id: str,
    req: UpsertSkillsPayload,
    auth_payload: dict = Depends(get_current_user),
):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    
    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")

    instance = InstanceModel._getById(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")

    if instance.get("status") not in ["running", "stopped", "error"]:
        raise HTTPException(status_code=400, detail="Forbidden: Agent must be in 'running', 'stopped', or 'error' status to perform skills updates.")
        
    if instance.get("status") == "running":
        InstanceModel._updateStatus(instance_id, "updating")

    from fastapp.services.agentProfileService import DEFAULT_SKILL_CONTENTS, SYSTEM_SKILLS_SET, build_skill_snapshot
    from bson import ObjectId
    from fastapp.database import get_db

    # Safeguard: Absolutely make sure system-managed skills cannot be edited manually or removed
    # Find which system skills should be active/enabled on this instance
    REQUIRED_SKILLS = {"bihand", "bihand-agent", "bihand-browser-use"}
    enabled_skills = set(instance.get("enabledSkills", []) or [])
    should_have_system_skills = REQUIRED_SKILLS.union(enabled_skills.intersection(SYSTEM_SKILLS_SET))

    # Incoming files mapped by name
    incoming_dict = {f.name: f.content for f in req.files if f.name.strip()}

    # Merge/Restore system skills to their exact original template contents
    final_files_dict = {}
    
    # 1. First add any custom skills sent by the user
    for name, content in incoming_dict.items():
        if name not in SYSTEM_SKILLS_SET:
            final_files_dict[name] = content

    # 2. Add or overwrite all required/enabled system skills with their official, unmodified contents
    for name in should_have_system_skills:
        final_files_dict[name] = DEFAULT_SKILL_CONTENTS[name]

    # Structure the finalized list of files
    final_files = [{"name": name, "content": content} for name, content in final_files_dict.items()]

    # 3. Save to MongoDB (skills backup sync recover)
    get_db()["instances"].update_one(
        {"_id": ObjectId(instance_id)},
        {"$set": {"skillsFiles": final_files, "updatedDate": datetime.now(timezone.utc)}}
    )

    # After setting the DB, trigger the background configuration push
    from fastapp.tasks import push_agent_config_task
    push_agent_config_task.delay(
        instance_id,
        instance.get("agentMd", ""),
        instance.get("soulMd", ""),
        instance.get("toolsMd", ""),
        instance.get("mcpConfig", ""),
        instance.get("enabledSkills", []),
    )
            
    return {"success": True, "files": final_files}

@fleetRouter.post("/{fleet_id}/instances/{instance_id}/skills/sync", summary="Sync desired skills")
async def sync_agent_skills(
    fleet_id: str,
    instance_id: str,
    req: AgentSkillSyncPayload,
    auth_payload: dict = Depends(get_current_user),
):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    
    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")

    instance = InstanceModel._getByIdWithKeys(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")

    if instance.get("status") != "running":
        raise HTTPException(status_code=400, detail="Forbidden: Agent must be in 'running' status to perform skills updates.")
        
    InstanceModel._updateStatus(instance_id, "updating")

    next_adapter_config, desired_skills, _ = sync_skills(instance, req.desiredSkills or [])
    InstanceModel._setAdapterConfig(instance_id, next_adapter_config)

    enabled_skill_slugs = [
        skill.split("/")[-1]
        for skill in desired_skills
        if isinstance(skill, str) and skill.strip() and skill.split("/")[-1] != "bihand"
    ]
    enabled_skill_slugs = list(dict.fromkeys(enabled_skill_slugs))
    InstanceModel._setEnabledSkills(instance_id, enabled_skill_slugs)

    # Trigger VM-level configuration and skill sync asynchronously.
    # On success, this background task will restore status to 'running'.
    from fastapp.tasks import push_agent_config_task
    push_agent_config_task.delay(
        instance_id,
        instance.get("agentMd", ""),
        instance.get("soulMd", ""),
        instance.get("toolsMd", ""),
        instance.get("mcpConfig", ""),
        enabled_skill_slugs,
    )

    updated_instance = InstanceModel._getById(instance_id)
    return build_skill_snapshot(updated_instance or instance)


@fleetRouter.get("/{fleet_id}/instances/{instance_id}/instructions", summary="Get instructions list")
async def get_agent_instructions(fleet_id: str, instance_id: str, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    
    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")

    instance = InstanceModel._getById(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")

    ip = instance.get("externalIp")
    from fastapp.services.provisioning import get_provisioning_strategy
    agent_type = instance.get("iteration", "openclaw")
    strategy = get_provisioning_strategy(agent_type)
    
    if instance.get("status") == "running" and ip:
        private_key = instance["sshKeyPrivate"]
        try:
            vm_files = strategy.getInstructions(ip, private_key)
            if vm_files and any(f.get("content", "").strip() for f in vm_files):
                # Sync back/backup instructions to DB cleanly
                try:
                    from bson import ObjectId
                    from fastapp.database import get_db
                    update_fields = {
                        "instructionsFiles": vm_files,
                        "updatedDate": datetime.now(timezone.utc)
                    }
                    
                    # Update individual fields for backwards-compatibility
                    for f in vm_files:
                        name = f.get("name")
                        content = f.get("content")
                        if name in ["AGENTS.md", "IDENTITY.md", "CLAUDE.md"]:
                            update_fields["agentMd"] = content
                        elif name == "SOUL.md":
                            update_fields["soulMd"] = content
                        elif name == "TOOLS.md":
                            update_fields["toolsMd"] = content
                        elif name == "HEARTBEAT.md":
                            update_fields["heartbeatMd"] = content

                    get_db()["instances"].update_one(
                        {"_id": ObjectId(instance_id)},
                        {"$set": update_fields}
                    )
                except Exception as db_e:
                    logger.error(f"Failed to backup VM instructions to MongoDB: {db_e}")
                
                # Replace with clean custom instructions in response payload for UI editing
                custom_instructions = instance.get("customAgentMd") or ""
                for f in vm_files:
                    if f.get("name") in ["AGENTS.md", "IDENTITY.md", "CLAUDE.md"]:
                        f["content"] = custom_instructions
                return {"files": vm_files}
            else:
                raise Exception("Fetched instructions are empty or VM directory not initialized yet")
        except Exception:
            pass

    # Fallback: Retrieve instructions from MongoDB backup if available
    db_instructions = instance.get("instructionsFiles")
    if db_instructions:
        # Load user custom instructions from DB only so system rules remain completely hidden in the UI
        custom_instructions = instance.get("customAgentMd") or ""
        for f in db_instructions:
            if f.get("name") in ["AGENTS.md", "IDENTITY.md", "CLAUDE.md"]:
                f["content"] = custom_instructions
        return {"files": db_instructions}

    # fallback using strategy-driven default files
    fallback_files = strategy.fallback_instructions(instance)
    custom_instructions = instance.get("customAgentMd") or ""
    for f in fallback_files:
        if f.get("name") in ["AGENTS.md", "IDENTITY.md", "CLAUDE.md"]:
            f["content"] = custom_instructions
    return {"files": fallback_files}

class InstructionFilePayload(BaseModel):
    name: str
    content: str

class UpsertInstructionsPayload(BaseModel):
    files: List[InstructionFilePayload]

@fleetRouter.put("/{fleet_id}/instances/{instance_id}/instructions", summary="Update instructions list")
async def put_agent_instructions_list(
    fleet_id: str,
    instance_id: str,
    req: UpsertInstructionsPayload,
    auth_payload: dict = Depends(get_current_user),
):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    
    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")

    instance = InstanceModel._getById(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")

    if instance.get("status") not in ["running", "stopped", "error"]:
        raise HTTPException(status_code=400, detail="Forbidden: Agent must be in 'running', 'stopped', or 'error' status to perform instructions updates.")
        
    if instance.get("status") == "running":
        InstanceModel._updateStatus(instance_id, "updating")

    from fastapp.services.provisioning import get_provisioning_strategy
    agent_type = instance.get("iteration", "openclaw")
    strategy = get_provisioning_strategy(agent_type)

    # Sanitize and strictly enforce critical system-managed instructions (DEFAULT_AGENT_MD) are append-only protected
    from fastapp.services.agentProfileService import DEFAULT_AGENT_MD
    sanitized_files = []
    user_custom_instructions = ""
    for f in req.files:
        name = f.name
        content = f.content
        if name in ["AGENTS.md", "IDENTITY.md", "CLAUDE.md"]:
            user_custom_instructions = content
            # Ensure the execution contract rules cannot be wiped out by prepending the master contract back automatically
            cleaned_master = DEFAULT_AGENT_MD.strip()
            if cleaned_master not in content:
                content = f"{cleaned_master}\n\n# User Instructions Override\n\n{content}"
        sanitized_files.append({"name": name, "content": content})

    # Update DB cache dynamically using strategy mapping
    strategy.update_db_from_instructions(instance_id, sanitized_files)

    # Save the full list of instruction files (name/path and content) to MongoDB for sync and recovery
    from bson import ObjectId
    from fastapp.database import get_db
    try:
        get_db()["instances"].update_one(
            {"_id": ObjectId(instance_id)},
            {"$set": {
                "customAgentMd": user_custom_instructions,
                "instructionsFiles": sanitized_files,
                "updatedDate": datetime.now(timezone.utc)
            }}
        )
    except Exception as db_e:
        logger.error(f"Failed to save instructions list backup to MongoDB: {db_e}")

    # Save to DB first, then trigger push_agent_config_task in background to sync to VM
    from fastapp.tasks import push_agent_config_task
    push_agent_config_task.delay(
        instance_id,
        instance.get("agentMd", ""),
        instance.get("soulMd", ""),
        instance.get("toolsMd", ""),
        instance.get("mcpConfig", ""),
        instance.get("enabledSkills", []),
    )

    return {"success": True}


class ConnectSocialMediaPayload(BaseModel):
    platform: str = Field(..., description="e.g. 'facebook', 'instagram', 'x', 'reddit'")
    credentialId: str = Field(..., description="The ID of the social media credential to bind to this agent")


@fleetRouter.get("/{fleet_id}/instances/{instance_id}/tools", summary="List agent tool integrations")
async def list_instance_tools(fleet_id: str, instance_id: str, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    
    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")

    instance = InstanceModel._getById(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")

    tool_conns = _sanitized_tool_connections(instance)
    
    # Platform-specific details mapping. Facebook posting is now handled via Meta MCP
    # (see META_MCP_TOOL_KEY above) instead of this legacy per-platform connect flow -
    # the /api/internal/social/post endpoint and post_to_facebook() stay intact so any
    # already-provisioned VM still running the old "bihand post facebook" skill keeps working.
    social_platforms = ["instagram", "x", "reddit"]
    social_tools_list = []
    
    for pf in social_platforms:
        pf_details = tool_conns.get(f"social_{pf}") or {}
        # Support fallback check to DB keys if any
        bound_id = (instance.get("socialCredentials") or {}).get(pf)
        if bound_id and pf_details.get("status") != "connected":
            pf_details = {
                "status": "connected",
                "credentialId": bound_id
            }
        
        social_tools_list.append({
            "key": f"social_{pf}",
            "label": f"Social Media: {pf.capitalize() if pf != 'x' else 'X (Twitter)'}",
            "status": pf_details.get("status", "not_connected"),
            "details": pf_details,
        })

    return {
        "tools": [
            {
                "key": GOOGLE_TOOL_KEY,
                "label": "Google Workspace",
                "status": tool_conns.get(GOOGLE_TOOL_KEY, {}).get("status", "not_connected"),
                "details": tool_conns.get(GOOGLE_TOOL_KEY, {}),
            },
            {
                "key": META_MCP_TOOL_KEY,
                "label": "Meta MCP (Facebook/Instagram/Threads/Ads Agent Tools)",
                "status": tool_conns.get(META_MCP_TOOL_KEY, {}).get("status", "not_connected"),
                "details": tool_conns.get(META_MCP_TOOL_KEY, {}),
            },
            *social_tools_list
        ]
    }


@fleetRouter.post("/{fleet_id}/instances/{instance_id}/tools/social-media/connect", summary="Connect/Bind Social Media Credential to an agent")
async def connect_social_media(
    fleet_id: str,
    instance_id: str,
    req: ConnectSocialMediaPayload,
    auth_payload: dict = Depends(get_current_user)
):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    
    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")

    instance = InstanceModel._getById(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")

    if instance.get("status") != "running":
        raise HTTPException(status_code=400, detail="Forbidden: Agent must be in 'running' status to perform updates.")

    InstanceModel._updateStatus(instance_id, "updating")

    platform_lower = req.platform.lower()

    # Verify credential exists and belongs to the user
    from fastapp.models.credentialModel import CredentialModel
    cred = CredentialModel.get_by_id(req.credentialId)
    if not cred or (cred.get("userId") != email and user_role != "admin"):
        raise HTTPException(status_code=404, detail="Social media credential not found or access denied")

    # Set specific platform credentialId in the instance
    InstanceModel._setPlatformCredentialId(instance_id, platform_lower, req.credentialId)

    # Set tool connection status
    InstanceModel._setToolConnection(instance_id, f"social_{platform_lower}", {
        "status": "connected",
        "connectedAt": datetime.now(timezone.utc).isoformat(),
        "credentialId": req.credentialId,
        "name": cred.get("name"),
    })

    # Enable the platform-specific social skill if not already enabled
    skill_slug = f"social-{platform_lower}"
    enabled_skills = instance.get("enabledSkills", [])
    if skill_slug not in enabled_skills:
        enabled_skills.append(skill_slug)
        InstanceModel._setEnabledSkills(instance_id, enabled_skills)
        next_adapter_config, _, _ = sync_skills(instance, enabled_skills)
        InstanceModel._setAdapterConfig(instance_id, next_adapter_config)
        
        # Trigger sync task to VM
        from fastapp.tasks import setup_social_media_tool_task
        setup_social_media_tool_task.delay(instance_id)

    # Trigger VM config push which transitions status back to running on complete
    from fastapp.tasks import push_agent_config_task
    push_agent_config_task.delay(
        instance_id,
        instance.get("agentMd", ""),
        instance.get("soulMd", ""),
        instance.get("toolsMd", ""),
        instance.get("mcpConfig", ""),
        enabled_skills,
    )

    return {"message": f"{platform_lower.capitalize()} connected successfully", "credentialId": req.credentialId}


@fleetRouter.post("/{fleet_id}/instances/{instance_id}/tools/social-media/disconnect", summary="Disconnect/Unbind Social Media Credential from an agent")
async def disconnect_social_media(
    fleet_id: str,
    instance_id: str,
    req: ConnectSocialMediaPayload, # Reuse model to specify which platform to disconnect
    auth_payload: dict = Depends(get_current_user)
):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    
    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")

    instance = InstanceModel._getById(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")

    if instance.get("status") != "running":
        raise HTTPException(status_code=400, detail="Forbidden: Agent must be in 'running' status to perform updates.")

    InstanceModel._updateStatus(instance_id, "updating")

    platform_lower = req.platform.lower()

    # Unbind specific platform credentialId in the instance
    InstanceModel._setPlatformCredentialId(instance_id, platform_lower, None)

    # Set tool connection status
    InstanceModel._setToolConnection(instance_id, f"social_{platform_lower}", {
        "status": "not_connected",
        "connectedAt": None,
        "credentialId": None,
        "name": None,
        "updatedAt": datetime.now(timezone.utc).isoformat(),
    })

    # Disable the platform-specific social skill
    skill_slug = f"social-{platform_lower}"
    enabled_skills = [s for s in instance.get("enabledSkills", []) if s != skill_slug]
    InstanceModel._setEnabledSkills(instance_id, enabled_skills)

    next_adapter_config, _, _ = sync_skills(instance, enabled_skills)
    InstanceModel._setAdapterConfig(instance_id, next_adapter_config)

    # Trigger VM config push which transitions status back to running on complete
    from fastapp.tasks import push_agent_config_task
    push_agent_config_task.delay(
        instance_id,
        instance.get("agentMd", ""),
        instance.get("soulMd", ""),
        instance.get("toolsMd", ""),
        instance.get("mcpConfig", ""),
        enabled_skills,
    )

    return {"message": f"{platform_lower.capitalize()} credential disconnected successfully"}


def _merge_meta_mcp_server(mcp_config_str: str) -> str:
    """Adds/overwrites the Facebook/Instagram/Threads/Ads MCP server entry in an mcpConfig
    JSON string. The access token is a ${cred:meta_mcp} placeholder, resolved only at
    SSH-push time (see fastapp/utils/mcpCredentials.py) - never stored or shown resolved."""
    import json
    try:
        parsed = json.loads(mcp_config_str) if mcp_config_str else {}
        if not isinstance(parsed, dict):
            parsed = {}
    except (TypeError, ValueError):
        parsed = {}

    if not isinstance(parsed.get("mcpServers"), dict):
        parsed["mcpServers"] = {}

    parsed["mcpServers"][META_MCP_SERVER_NAME] = {
        "command": "npx",
        "args": ["-y", "@oliverames/meta-mcp-server"],
        "env": {"META_ACCESS_TOKEN": f"${{cred:{META_MCP_TOOL_KEY}}}"},
    }
    return json.dumps(parsed)


def _remove_meta_mcp_server(mcp_config_str: str) -> str:
    import json
    try:
        parsed = json.loads(mcp_config_str) if mcp_config_str else {}
    except (TypeError, ValueError):
        return mcp_config_str

    if isinstance(parsed, dict) and isinstance(parsed.get("mcpServers"), dict):
        parsed["mcpServers"].pop(META_MCP_SERVER_NAME, None)
    return json.dumps(parsed)


class ConnectMetaMcpPayload(BaseModel):
    credentialId: str = Field(..., description="Facebook Page / System User access-token credential to use for Meta MCP tools")


@fleetRouter.post("/{fleet_id}/instances/{instance_id}/tools/meta-mcp/connect", summary="Connect Meta MCP (Facebook/Instagram/Threads/Ads agent tools) to an agent")
async def connect_meta_mcp(
    fleet_id: str,
    instance_id: str,
    req: ConnectMetaMcpPayload,
    auth_payload: dict = Depends(get_current_user)
):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"

    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")

    instance = InstanceModel._getById(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")

    if instance.get("status") != "running":
        raise HTTPException(status_code=400, detail="Forbidden: Agent must be in 'running' status to perform updates.")

    from fastapp.models.credentialModel import CredentialModel
    cred = CredentialModel.get_by_id(req.credentialId)
    if not cred or (cred.get("userId") != email and user_role != "admin"):
        raise HTTPException(status_code=404, detail="Credential not found or access denied")

    InstanceModel._updateStatus(instance_id, "updating")

    InstanceModel._setToolConnection(instance_id, META_MCP_TOOL_KEY, {
        "status": "connected",
        "connectedAt": datetime.now(timezone.utc).isoformat(),
        "credentialId": req.credentialId,
        "name": cred.get("name"),
    })

    merged_mcp_config = _merge_meta_mcp_server(instance.get("mcpConfig", ""))
    InstanceModel._updateConfig(instance_id, None, None, None, merged_mcp_config)

    # Enable the meta-mcp skill (tells the agent to use the native MCP tools) if not already
    enabled_skills = instance.get("enabledSkills", []) or []
    if META_MCP_SKILL_SLUG not in enabled_skills:
        enabled_skills = enabled_skills + [META_MCP_SKILL_SLUG]
        InstanceModel._setEnabledSkills(instance_id, enabled_skills)
        next_adapter_config, _, _ = sync_skills(instance, enabled_skills)
        InstanceModel._setAdapterConfig(instance_id, next_adapter_config)

    from fastapp.tasks import push_agent_config_task
    push_agent_config_task.delay(
        instance_id,
        instance.get("agentMd", ""),
        instance.get("soulMd", ""),
        instance.get("toolsMd", ""),
        merged_mcp_config,
        enabled_skills,
    )

    return {"message": "Meta MCP connected successfully", "credentialId": req.credentialId}


@fleetRouter.post("/{fleet_id}/instances/{instance_id}/tools/meta-mcp/disconnect", summary="Disconnect Meta MCP from an agent")
async def disconnect_meta_mcp(fleet_id: str, instance_id: str, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"

    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")

    instance = InstanceModel._getById(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")

    if instance.get("status") != "running":
        raise HTTPException(status_code=400, detail="Forbidden: Agent must be in 'running' status to perform updates.")

    InstanceModel._updateStatus(instance_id, "updating")

    InstanceModel._setToolConnection(instance_id, META_MCP_TOOL_KEY, {
        "status": "not_connected",
        "connectedAt": None,
        "credentialId": None,
        "name": None,
        "updatedAt": datetime.now(timezone.utc).isoformat(),
    })

    remaining_mcp_config = _remove_meta_mcp_server(instance.get("mcpConfig", ""))
    InstanceModel._updateConfig(instance_id, None, None, None, remaining_mcp_config)

    enabled_skills = [s for s in instance.get("enabledSkills", []) if s != META_MCP_SKILL_SLUG]
    InstanceModel._setEnabledSkills(instance_id, enabled_skills)
    next_adapter_config, _, _ = sync_skills(instance, enabled_skills)
    InstanceModel._setAdapterConfig(instance_id, next_adapter_config)

    from fastapp.tasks import push_agent_config_task
    push_agent_config_task.delay(
        instance_id,
        instance.get("agentMd", ""),
        instance.get("soulMd", ""),
        instance.get("toolsMd", ""),
        remaining_mcp_config,
        enabled_skills,
    )

    return {"message": "Meta MCP disconnected successfully"}


def _merge_meta_devtools_mcp_server(mcp_config_str: str) -> str:
    """Adds/overwrites Meta's remote Developer Tools MCP server entry in an mcpConfig JSON
    string. The bearer token is a ${cred:meta_devtools} placeholder, resolved only at
    SSH-push time (see fastapp/utils/mcpCredentials.py) - never stored or shown resolved."""
    import json
    try:
        parsed = json.loads(mcp_config_str) if mcp_config_str else {}
        if not isinstance(parsed, dict):
            parsed = {}
    except (TypeError, ValueError):
        parsed = {}

    if not isinstance(parsed.get("mcpServers"), dict):
        parsed["mcpServers"] = {}

    parsed["mcpServers"][META_DEVTOOLS_SERVER_NAME] = {
        "url": "https://mcp.facebook.com/devtools",
        "headers": {"Authorization": f"Bearer ${{cred:{META_DEVTOOLS_TOOL_KEY}}}"},
    }
    return json.dumps(parsed)


def _remove_meta_devtools_mcp_server(mcp_config_str: str) -> str:
    import json
    try:
        parsed = json.loads(mcp_config_str) if mcp_config_str else {}
    except (TypeError, ValueError):
        return mcp_config_str

    if isinstance(parsed, dict) and isinstance(parsed.get("mcpServers"), dict):
        parsed["mcpServers"].pop(META_DEVTOOLS_SERVER_NAME, None)
    return json.dumps(parsed)


class ConnectMetaDevtoolsMcpPayload(BaseModel):
    credentialId: str = Field(..., description="Meta OAuth credential to use for the Developer Tools MCP")


@fleetRouter.post("/{fleet_id}/instances/{instance_id}/tools/meta-devtools/connect", summary="Connect Meta Developer Tools MCP to an agent")
async def connect_meta_devtools_mcp(
    fleet_id: str,
    instance_id: str,
    req: ConnectMetaDevtoolsMcpPayload,
    auth_payload: dict = Depends(get_current_user)
):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"

    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")

    instance = InstanceModel._getById(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")

    if instance.get("status") != "running":
        raise HTTPException(status_code=400, detail="Forbidden: Agent must be in 'running' status to perform updates.")

    from fastapp.models.credentialModel import CredentialModel
    cred = CredentialModel.get_by_id(req.credentialId)
    if not cred or (cred.get("userId") != email and user_role != "admin"):
        raise HTTPException(status_code=404, detail="Credential not found or access denied")

    InstanceModel._updateStatus(instance_id, "updating")

    InstanceModel._setToolConnection(instance_id, META_DEVTOOLS_TOOL_KEY, {
        "status": "connected",
        "connectedAt": datetime.now(timezone.utc).isoformat(),
        "credentialId": req.credentialId,
        "name": cred.get("name"),
    })

    merged_mcp_config = _merge_meta_devtools_mcp_server(instance.get("mcpConfig", ""))
    InstanceModel._updateConfig(instance_id, None, None, None, merged_mcp_config)

    from fastapp.tasks import push_agent_config_task
    push_agent_config_task.delay(
        instance_id,
        instance.get("agentMd", ""),
        instance.get("soulMd", ""),
        instance.get("toolsMd", ""),
        merged_mcp_config,
        instance.get("enabledSkills", []),
    )

    return {"message": "Meta Developer Tools MCP connected successfully", "credentialId": req.credentialId}


@fleetRouter.post("/{fleet_id}/instances/{instance_id}/tools/meta-devtools/disconnect", summary="Disconnect Meta Developer Tools MCP from an agent")
async def disconnect_meta_devtools_mcp(fleet_id: str, instance_id: str, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"

    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")

    instance = InstanceModel._getById(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")

    if instance.get("status") != "running":
        raise HTTPException(status_code=400, detail="Forbidden: Agent must be in 'running' status to perform updates.")

    InstanceModel._updateStatus(instance_id, "updating")

    InstanceModel._setToolConnection(instance_id, META_DEVTOOLS_TOOL_KEY, {
        "status": "not_connected",
        "connectedAt": None,
        "credentialId": None,
        "name": None,
        "updatedAt": datetime.now(timezone.utc).isoformat(),
    })

    remaining_mcp_config = _remove_meta_devtools_mcp_server(instance.get("mcpConfig", ""))
    InstanceModel._updateConfig(instance_id, None, None, None, remaining_mcp_config)

    from fastapp.tasks import push_agent_config_task
    push_agent_config_task.delay(
        instance_id,
        instance.get("agentMd", ""),
        instance.get("soulMd", ""),
        instance.get("toolsMd", ""),
        remaining_mcp_config,
        instance.get("enabledSkills", []),
    )

    return {"message": "Meta Developer Tools MCP disconnected successfully"}


# --- Customer support flows: fleet-owned (not instance-owned) channel connections + policy.
# Agents are *assigned* to operate a flow (assignedInstanceId) and that can be reassigned;
# see fastapp/models/flowModel.py for the full ownership/permission model, mirrored here for
# the human (JWT) side - the fleet owner always has full CRUD, same as every other resource
# in this file. Agents get an M2M-authenticated equivalent in agentM2MController.py, gated by
# the root-agent-by-default / explicit-grant permission model instead of JWT ownership.

def _require_fleet_owner(fleet_id: str, auth_payload: dict) -> dict:
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"

    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")
    return fleet


class StageDefPayload(BaseModel):
    key: str = Field(..., description="Stable identifier, referenced by Conversation.currentStageKey")
    name: str
    goal: str = Field(default="", description="What the agent should accomplish in this stage")
    exitCriteria: str = Field(default="", description="What must be true before the model marks stage_complete")
    escalateToHuman: bool = Field(default=False, description="Completing this stage forces human review regardless of the flow's mode")
    maxTurns: Optional[int] = Field(default=None, description="Flag for human review if the conversation stays in this stage past this many turns")


def _validate_stages(stages: Optional[List[Dict[str, Any]]]) -> None:
    if not stages:
        return
    keys = [s.get("key") for s in stages]
    if any(not k for k in keys):
        raise HTTPException(status_code=400, detail="Every stage requires a non-empty key")
    if len(keys) != len(set(keys)):
        raise HTTPException(status_code=400, detail="Stage keys must be unique within a flow")


class FlowCreatePayload(BaseModel):
    name: str
    platform: str = Field(..., description="'messenger' or 'zalo'")
    channelType: str = Field(..., description="'page_webhook', 'oa_webhook', or 'personal_browser'")
    assignedInstanceId: Optional[str] = Field(default=None, description="Which agent operates this flow - required for personal_browser (installs channel_sync.py on that agent's VM)")
    credentialId: Optional[str] = Field(default=None, description="Business tiers: Facebook Page / Zalo OA access-token credential")
    pageId: Optional[str] = None
    oaId: Optional[str] = None
    verifyToken: Optional[str] = None
    label: Optional[str] = Field(default=None, description="Personal tiers: human-friendly label for which account this is")
    stages: Optional[List[StageDefPayload]] = Field(default=None, description="Optional ordered funnel - omit for a flow that responds without stage tracking")


@fleetRouter.get("/{fleet_id}/flows", summary="List customer-support flows for a fleet")
async def list_flows(fleet_id: str, auth_payload: dict = Depends(get_current_user)):
    _require_fleet_owner(fleet_id, auth_payload)
    from fastapp.models.flowModel import FlowModel
    return {"flows": FlowModel._listByFleet(fleet_id)}


@fleetRouter.post("/{fleet_id}/flows", summary="Create a customer-support flow (connect a channel)")
async def create_flow(fleet_id: str, req: FlowCreatePayload, auth_payload: dict = Depends(get_current_user)):
    _require_fleet_owner(fleet_id, auth_payload)

    if req.platform not in ("messenger", "zalo"):
        raise HTTPException(status_code=400, detail="platform must be 'messenger' or 'zalo'")
    if req.channelType not in ("page_webhook", "oa_webhook", "personal_browser"):
        raise HTTPException(status_code=400, detail="channelType must be 'page_webhook', 'oa_webhook', or 'personal_browser'")

    credential_name = None
    if req.credentialId:
        from fastapp.models.credentialModel import CredentialModel
        email = auth_payload["email"]
        user_role = auth_payload.get("role", "user")
        from fastapp.utils.adminAuth import ADMIN_EMAILS
        if auth_payload.get("email") in ADMIN_EMAILS:
            user_role = "admin"
        cred = CredentialModel.get_by_id(req.credentialId)
        if not cred or (cred.get("userId") != email and user_role != "admin"):
            raise HTTPException(status_code=404, detail="Credential not found or access denied")
        credential_name = cred.get("name")

    if req.channelType == "personal_browser":
        if not req.assignedInstanceId:
            raise HTTPException(status_code=400, detail="assignedInstanceId is required for personal_browser flows")
        instance = InstanceModel._getById(req.assignedInstanceId)
        if not instance or instance.get("fleetId") != fleet_id:
            raise HTTPException(status_code=404, detail="Assigned instance not found")
        if instance.get("status") != "running":
            raise HTTPException(status_code=400, detail="Assigned agent must be in 'running' status to install the channel sync script.")

    stages_dicts = [s.dict() for s in req.stages] if req.stages else None
    _validate_stages(stages_dicts)

    from fastapp.models.flowModel import FlowModel
    flow = FlowModel._create(
        fleet_id=fleet_id,
        name=req.name,
        platform=req.platform,
        channel_type=req.channelType,
        created_by=f"human:{auth_payload['email']}",
        assigned_instance_id=req.assignedInstanceId,
        page_id=req.pageId,
        oa_id=req.oaId,
        verify_token=req.verifyToken,
        credential_id=req.credentialId,
        label=req.label,
        stages=stages_dicts,
    )

    if req.channelType == "personal_browser":
        from fastapp.tasks import setup_personal_channel_sync_task
        setup_personal_channel_sync_task.delay(req.assignedInstanceId)

    if credential_name:
        flow["credentialName"] = credential_name
    return {"flow": flow}


@fleetRouter.get("/{fleet_id}/flows/{flow_id}", summary="Get a customer-support flow")
async def get_flow(fleet_id: str, flow_id: str, auth_payload: dict = Depends(get_current_user)):
    _require_fleet_owner(fleet_id, auth_payload)
    from fastapp.models.flowModel import FlowModel
    flow = FlowModel._getById(flow_id)
    if not flow or flow.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Flow not found")
    return {"flow": flow}


class FlowUpdatePayload(BaseModel):
    name: Optional[str] = None
    supportPolicy: Optional[Dict[str, Any]] = None
    status: Optional[str] = None
    stages: Optional[List[StageDefPayload]] = Field(default=None, description="Replaces the entire funnel - omit to leave stages unchanged, pass [] to clear it")


@fleetRouter.patch("/{fleet_id}/flows/{flow_id}", summary="Update a customer-support flow's name/policy/status/funnel")
async def update_flow(fleet_id: str, flow_id: str, req: FlowUpdatePayload, auth_payload: dict = Depends(get_current_user)):
    _require_fleet_owner(fleet_id, auth_payload)
    from fastapp.models.flowModel import FlowModel
    flow = FlowModel._getById(flow_id)
    if not flow or flow.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Flow not found")

    updates = {k: v for k, v in req.dict(exclude_unset=True).items() if v is not None}
    if "stages" in updates:
        updates["stages"] = [s if isinstance(s, dict) else s.dict() for s in updates["stages"]]
        _validate_stages(updates["stages"])
    FlowModel._update(flow_id, updates)
    return {"flow": FlowModel._getById(flow_id)}


@fleetRouter.delete("/{fleet_id}/flows/{flow_id}", summary="Delete a customer-support flow")
async def delete_flow(fleet_id: str, flow_id: str, auth_payload: dict = Depends(get_current_user)):
    _require_fleet_owner(fleet_id, auth_payload)
    from fastapp.models.flowModel import FlowModel
    flow = FlowModel._getById(flow_id)
    if not flow or flow.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Flow not found")
    FlowModel._delete(flow_id)
    return {"message": "Flow deleted"}


class FlowReassignPayload(BaseModel):
    instanceId: str = Field(..., description="Agent to hand this flow's conversations to - takes effect immediately for existing conversations, not just new ones")


@fleetRouter.post("/{fleet_id}/flows/{flow_id}/reassign", summary="Reassign which agent operates a flow")
async def reassign_flow(fleet_id: str, flow_id: str, req: FlowReassignPayload, auth_payload: dict = Depends(get_current_user)):
    _require_fleet_owner(fleet_id, auth_payload)
    from fastapp.models.flowModel import FlowModel
    flow = FlowModel._getById(flow_id)
    if not flow or flow.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Flow not found")

    instance = InstanceModel._getById(req.instanceId)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Target instance not found")

    FlowModel._update(flow_id, {"assignedInstanceId": req.instanceId})

    if flow.get("channelType") == "personal_browser" and instance.get("status") == "running":
        from fastapp.tasks import setup_personal_channel_sync_task
        setup_personal_channel_sync_task.delay(req.instanceId)

    return {"flow": FlowModel._getById(flow_id)}


class FlowAccessGrantPayload(BaseModel):
    instanceId: str
    role: str = Field(default="viewer", description="'viewer', 'editor', or 'owner'")


@fleetRouter.post("/{fleet_id}/flows/{flow_id}/access", summary="Grant an agent access to a flow (human-only)")
async def grant_flow_access(fleet_id: str, flow_id: str, req: FlowAccessGrantPayload, auth_payload: dict = Depends(get_current_user)):
    _require_fleet_owner(fleet_id, auth_payload)
    from fastapp.models.flowModel import FlowModel
    flow = FlowModel._getById(flow_id)
    if not flow or flow.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Flow not found")
    if req.role not in ("viewer", "editor", "owner"):
        raise HTTPException(status_code=400, detail="role must be 'viewer', 'editor', or 'owner'")

    instance = InstanceModel._getById(req.instanceId)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")

    FlowModel._grantAccess(flow_id, req.instanceId, req.role)
    return {"flow": FlowModel._getById(flow_id)}


@fleetRouter.delete("/{fleet_id}/flows/{flow_id}/access/{instance_id}", summary="Revoke an agent's access to a flow (human-only)")
async def revoke_flow_access(fleet_id: str, flow_id: str, instance_id: str, auth_payload: dict = Depends(get_current_user)):
    _require_fleet_owner(fleet_id, auth_payload)
    from fastapp.models.flowModel import FlowModel
    flow = FlowModel._getById(flow_id)
    if not flow or flow.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Flow not found")
    FlowModel._revokeAccess(flow_id, instance_id)
    return {"flow": FlowModel._getById(flow_id)}


class McpServerPayload(BaseModel):
    name: str = Field(..., description="Unique MCP server identifier")
    command: Optional[str] = Field(default=None, description="Executable for a local/stdio MCP server")
    args: Optional[List[str]] = Field(default_factory=list, description="Arguments for a local/stdio MCP server")
    env: Optional[Dict[str, str]] = Field(default_factory=dict, description="Environment variables for a local/stdio MCP server")
    url: Optional[str] = Field(default=None, description="Endpoint URL for a remote MCP server")
    headers: Optional[Dict[str, str]] = Field(default_factory=dict, description="Headers for a remote MCP server")


def _mcp_server_public_shape(name: str, config: Dict[str, Any]) -> Dict[str, Any]:
    is_remote = bool(config.get("url"))
    return {
        "name": name,
        "command": config.get("command") if not is_remote else None,
        "args": config.get("args", []) if not is_remote else [],
        "env": config.get("env", {}) if not is_remote else {},
        "url": config.get("url") if is_remote else None,
        "headers": config.get("headers", {}) if is_remote else {},
        "protected": name in PROTECTED_MCP_SERVER_NAMES,
    }


def _upsert_mcp_server(mcp_config_str: str, name: str, payload: McpServerPayload) -> str:
    """Adds/overwrites a single MCP server entry in an mcpConfig JSON string, tolerant of
    whichever of the 4 on-VM shapes the DB copy currently happens to be cached in (see
    mcp_normalizer.extract_all_servers). Always re-serializes to the standard mcpServers
    shape - per-agent-type normalization happens downstream at push time."""
    import json
    from fastapp.utils.mcp_normalizer import extract_all_servers
    servers = extract_all_servers(mcp_config_str)
    if payload.url:
        entry: Dict[str, Any] = {"url": payload.url}
        if payload.headers:
            entry["headers"] = payload.headers
    else:
        entry = {
            "command": payload.command or "",
            "args": payload.args or [],
            "env": payload.env or {},
        }
    servers[name] = entry
    return json.dumps({"mcpServers": servers}, indent=2)


def _remove_mcp_server_entry(mcp_config_str: str, name: str) -> str:
    import json
    from fastapp.utils.mcp_normalizer import extract_all_servers
    servers = extract_all_servers(mcp_config_str)
    servers.pop(name, None)
    return json.dumps({"mcpServers": servers}, indent=2)


def _require_running_instance_for_mcp(fleet_id: str, instance_id: str, auth_payload: dict):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"

    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")

    instance = InstanceModel._getById(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")

    if instance.get("status") != "running":
        raise HTTPException(status_code=400, detail="Forbidden: Agent must be in 'running' status to perform updates.")

    return instance


@fleetRouter.get("/{fleet_id}/instances/{instance_id}/mcp-servers", summary="List MCP servers configured for an agent")
async def list_mcp_servers(fleet_id: str, instance_id: str, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"

    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")

    instance = InstanceModel._getById(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")

    from fastapp.utils.mcp_normalizer import extract_all_servers
    servers = extract_all_servers(instance.get("mcpConfig", ""))
    return {"servers": [_mcp_server_public_shape(name, config) for name, config in servers.items()]}


@fleetRouter.post("/{fleet_id}/instances/{instance_id}/mcp-servers", summary="Add a custom MCP server to an agent")
async def add_mcp_server(fleet_id: str, instance_id: str, req: McpServerPayload, auth_payload: dict = Depends(get_current_user)):
    instance = _require_running_instance_for_mcp(fleet_id, instance_id, auth_payload)

    name = req.name.strip()
    if not name:
        raise HTTPException(status_code=400, detail="MCP server name is required")
    if name in PROTECTED_MCP_SERVER_NAMES:
        raise HTTPException(status_code=400, detail=f"'{name}' is a system-managed MCP server and cannot be modified")

    from fastapp.utils.mcp_normalizer import extract_all_servers
    existing = extract_all_servers(instance.get("mcpConfig", ""))
    if name in existing:
        raise HTTPException(status_code=409, detail=f"An MCP server named '{name}' already exists. Use edit instead.")

    new_config = _upsert_mcp_server(instance.get("mcpConfig", ""), name, req)
    InstanceModel._updateConfig(instance_id, None, None, None, new_config)
    InstanceModel._updateStatus(instance_id, "updating")

    from fastapp.tasks import push_agent_config_task
    push_agent_config_task.delay(
        instance_id,
        instance.get("agentMd", ""),
        instance.get("soulMd", ""),
        instance.get("toolsMd", ""),
        new_config,
        instance.get("enabledSkills", []),
    )

    return {"server": _mcp_server_public_shape(name, extract_all_servers(new_config)[name])}


@fleetRouter.put("/{fleet_id}/instances/{instance_id}/mcp-servers/{server_name}", summary="Edit a custom MCP server on an agent")
async def update_mcp_server(fleet_id: str, instance_id: str, server_name: str, req: McpServerPayload, auth_payload: dict = Depends(get_current_user)):
    instance = _require_running_instance_for_mcp(fleet_id, instance_id, auth_payload)

    if server_name in PROTECTED_MCP_SERVER_NAMES:
        raise HTTPException(status_code=400, detail=f"'{server_name}' is a system-managed MCP server and cannot be modified")

    from fastapp.utils.mcp_normalizer import extract_all_servers
    existing = extract_all_servers(instance.get("mcpConfig", ""))
    if server_name not in existing:
        raise HTTPException(status_code=404, detail=f"MCP server '{server_name}' not found")

    new_config = _upsert_mcp_server(instance.get("mcpConfig", ""), server_name, req)
    InstanceModel._updateConfig(instance_id, None, None, None, new_config)
    InstanceModel._updateStatus(instance_id, "updating")

    from fastapp.tasks import push_agent_config_task
    push_agent_config_task.delay(
        instance_id,
        instance.get("agentMd", ""),
        instance.get("soulMd", ""),
        instance.get("toolsMd", ""),
        new_config,
        instance.get("enabledSkills", []),
    )

    return {"server": _mcp_server_public_shape(server_name, extract_all_servers(new_config)[server_name])}


@fleetRouter.delete("/{fleet_id}/instances/{instance_id}/mcp-servers/{server_name}", summary="Remove a custom MCP server from an agent")
async def delete_mcp_server(fleet_id: str, instance_id: str, server_name: str, auth_payload: dict = Depends(get_current_user)):
    instance = _require_running_instance_for_mcp(fleet_id, instance_id, auth_payload)

    if server_name in PROTECTED_MCP_SERVER_NAMES:
        raise HTTPException(status_code=400, detail=f"'{server_name}' is a system-managed MCP server and cannot be removed")

    from fastapp.utils.mcp_normalizer import extract_all_servers
    existing = extract_all_servers(instance.get("mcpConfig", ""))
    if server_name not in existing:
        raise HTTPException(status_code=404, detail=f"MCP server '{server_name}' not found")

    new_config = _remove_mcp_server_entry(instance.get("mcpConfig", ""), server_name)
    InstanceModel._updateConfig(instance_id, None, None, None, new_config)
    InstanceModel._updateStatus(instance_id, "updating")

    from fastapp.tasks import push_agent_config_task
    push_agent_config_task.delay(
        instance_id,
        instance.get("agentMd", ""),
        instance.get("soulMd", ""),
        instance.get("toolsMd", ""),
        new_config,
        instance.get("enabledSkills", []),
    )

    return {"message": f"MCP server '{server_name}' removed successfully"}


@fleetRouter.post("/{fleet_id}/instances/{instance_id}/tools/google-workspace/disconnect", summary="Disconnect Google Workspace from an agent")
async def disconnect_google_workspace(fleet_id: str, instance_id: str, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    
    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")

    instance = InstanceModel._getById(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")

    if instance.get("status") != "running":
        raise HTTPException(status_code=400, detail="Forbidden: Agent must be in 'running' status to perform updates.")

    InstanceModel._updateStatus(instance_id, "updating")

    InstanceModel._setToolConnection(instance_id, GOOGLE_TOOL_KEY, {
        "status": "not_connected",
        "connectedAt": None,
        "oauthState": None,
        "email": None,
        "scopes": [],
        "credential": {},
        "updatedAt": datetime.now(timezone.utc),
    })

    enabled_skills = [s for s in instance.get("enabledSkills", []) if s != "bihand-google-workspace"]
    InstanceModel._setEnabledSkills(instance_id, enabled_skills)

    next_adapter_config, _, _ = sync_skills(instance, enabled_skills)
    InstanceModel._setAdapterConfig(instance_id, next_adapter_config)

    # Trigger VM config push which transitions status back to running on complete
    from fastapp.tasks import push_agent_config_task
    push_agent_config_task.delay(
        instance_id,
        instance.get("agentMd", ""),
        instance.get("soulMd", ""),
        instance.get("toolsMd", ""),
        instance.get("mcpConfig", ""),
        enabled_skills,
    )

    return {"message": "Google Workspace disconnected"}

@fleetRouter.delete("/{fleet_id}/instances/{instance_id}", summary="Destroy specific agent in a fleet")
async def delete_fleet_agent(fleet_id: str, instance_id: str, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    
    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")

    instance = InstanceModel._getById(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found in this fleet")
        
    # Re-parenting sub-agents: if this agent has sub-agents reporting to them,
    # promote them to report to this agent's manager (or None if this agent reports to no one)
    from fastapp.database import get_db
    from bson import ObjectId
    db = get_db()
    manager_id = instance.get("reportsTo")
    
    db["instances"].update_many(
        {"fleetId": fleet_id, "reportsTo": instance_id},
        {"$set": {"reportsTo": manager_id}}
    )

    InstanceModel._updateStatus(instance_id, "deleting")
    from fastapp.tasks import delete_instance_task
    delete_instance_task.delay(instance_id)
    
    return {"message": "Agent destruction initiated and subordinates re-parented"}

class UpdateAgentStructurePayload(BaseModel):
    reportsTo: Optional[str] = None
    role: Optional[str] = None
    title: Optional[str] = None

@fleetRouter.put("/{fleet_id}/instances/{instance_id}/structure", summary="Update Agent's reporting/role structure in a fleet")
async def update_agent_structure(fleet_id: str, instance_id: str, req: UpdateAgentStructurePayload, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    
    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")

    instance = InstanceModel._getById(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found in this fleet")
        
    # Validation: prevent cyclical reporting structure
    if req.reportsTo:
        if req.reportsTo == instance_id:
            raise HTTPException(status_code=400, detail="An agent cannot report to themselves.")
        
        # Walk up the chain to ensure no cycle
        from fastapp.database import get_db
        from bson import ObjectId
        db = get_db()
        current_manager = req.reportsTo
        visited = {instance_id}
        while current_manager:
            if current_manager in visited:
                raise HTTPException(status_code=400, detail="Cyclical reporting structure detected.")
            visited.add(current_manager)
            mgr_doc = db["instances"].find_one({"_id": ObjectId(current_manager)})
            current_manager = mgr_doc.get("reportsTo") if mgr_doc else None

    # Update database
    from fastapp.database import get_db
    from bson import ObjectId
    from datetime import datetime, timezone
    db = get_db()
    
    update_data = {}
    if req.reportsTo is not None:
        update_data["reportsTo"] = req.reportsTo if req.reportsTo else None
    if req.role is not None:
        update_data["fleetRole"] = req.role
        update_data["alias"] = f"{fleet['name']} - {req.role}"
    if req.title is not None:
        update_data["title"] = req.title
        
    if update_data:
        update_data["updatedDate"] = datetime.now(timezone.utc)
        db["instances"].update_one(
            {"_id": ObjectId(instance_id)},
            {"$set": update_data}
        )
        
    return {"message": "Agent structure updated successfully"}

class AddAgentRequest(BaseModel):
    agent: AgentConfig = Field(..., description="Agent configuration to add")

@fleetRouter.post("/{fleet_id}/instances", summary="Add a new agent to an existing fleet")
async def add_fleet_agent(fleet_id: str, req: AddAgentRequest, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    
    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")

    from fastapp.models.userModel import UserModel
    # Resolve the original fleet owner to correctly deduct upfront VM creation credits from their wallet
    owner_email = fleet["userId"]
    user = UserModel._getUserByEmail(owner_email)
    user_hash = user.get("hash", "") if user else ""

    ag = req.agent
    machine_credit_costs_per_day = {
        "e2-small": 100,
        "e2-medium": 200,
        "e2-standard-2": 400
    }
    mtype = ag.machineType or "e2-small"
    
    # Under daily pay-as-you-go model, we deduct the first day's credits upfront
    total_cost = machine_credit_costs_per_day.get(mtype, 100)

    current_credits = user.get("credits", 0) if user else 0
    if current_credits < total_cost:
        pass  # OSS build: no credit/billing gating (BYOK — bring your own GCP + LLM key)

    # Deduct credits
    tx_details = {
        "action": "add_agent_to_fleet",
        "fleetId": fleet_id,
        "fleetName": fleet.get("name"),
        "role": ag.role,
        "title": ag.title,
        "machineType": mtype,
        "iteration": ag.agentType
    }
    success = UserModel._deductCredits(owner_email, total_cost, details=tx_details)
    if not success:
         pass  # OSS build: no credit/billing gating (BYOK — bring your own GCP + LLM key)

    # Create instances in DB and queue provisioning task
    from fastapp.services import sshService
    from fastapp.controllers.instanceController import get_disk_size_gb
    import secrets

    role = ag.role
    title = ag.title or role
    agent_type = ag.agentType or "openclaw"
    provider = ag.provider or "openai"
    model = ag.model or "gpt-4"

    if provider.lower() == "bihand":
        model = "gemini-3.5-flash"
        ag.apiKey = "bihand-system-placeholder"

    # Generate SSH keys and instance metadata
    suffix = secrets.token_hex(4)
    vm_name = f"bh-{suffix}"
    disk_name = f"bh-disk-{suffix}"
    ssh_keys = sshService.generate_ssh_keypair()

    instance = InstanceModel._createInstance(
        userId=owner_email,
        userHash=user_hash,
        vmName=vm_name,
        zone="us-central1-a",
        machineType=mtype,
        diskName=disk_name,
        diskSizeGb=get_disk_size_gb(mtype),
        provider=provider,
        model=model,
        sshKeyPrivate=ssh_keys["private"],
        sshKeyPublic=ssh_keys["public"],
        createdBy=email,
        alias=f"{fleet['name']} - {role}",
        iteration=agent_type,
        fleetId=fleet_id,
        fleetRole=role,
        title=title,
        reportsTo=ag.reportsTo, # Real instance ID of manager
        agentMd=ag.agentMd if ag.agentMd else DEFAULT_AGENT_MD,
        customAgentMd=ag.customAgentMd or "",
        soulMd=ag.soulMd or "",
        heartbeatMd=ag.toolsMd or "",
        toolsMd=ag.toolsMd or "",
        mcpConfig=ag.mcpConfig or "",
        enabledSkills=ag.enabledSkills or [],
        avatarHash=ag.avatarHash,
        skillsFiles=ag.skillsFiles or [],
        oauthToken=ag.oauthToken or None,
        customBaseUrl=ag.customBaseUrl or None,
    )

    from fastapp.tasks import provision_instance_task
    provision_instance_task.delay(
        str(instance["_id"]),
        owner_email,
        provider,
        ag.apiKey, # Will be parsed / used
        "SantCorp123", # default password or master pass can be configured
        agent_type
    )

    # Append to fleet agents list
    from fastapp.database import get_db
    get_db()["fleets"].update_one(
        {"_id": fleet_id},
        {"$push": {"agents": ag.dict()}}
    )

    return {"message": "Agent addition initiated", "instanceId": str(instance["_id"])}

class InitialTaskConfig(BaseModel):
    title: str = Field(..., description="Task title")
    description: str = Field(..., description="Task description")

class ProvisionFleetRequest(BaseModel):
    name: str = Field(..., description="Name of the company/fleet")
    plan: str = Field(..., description="Plan tier: starter, medium, custom")
    agents: List[AgentConfig] = Field(default=[], description="Agent configurations")
    password: str = Field(..., description="Master password for the dashboard instances")
    apiBudget: float = Field(default=50.0, description="Monthly API Spend limit (USD)")
    mission: str = Field(default="Execute tasks autonomously.", description="Company Goal/Mission")
    initialTask: Optional[InitialTaskConfig] = Field(None, description="Initial task to create")

@fleetRouter.post("", summary="Provision a new AI Company Fleet")
async def provision_fleet(
    req: ProvisionFleetRequest,
    auth_payload: dict = Depends(get_current_user)
):
    """
    Provision a new company fleet (Bihand managed).
    Based on the plan, this spins up the necessary VMs and registers them under one fleet.
    """
    email = auth_payload["email"]
    
    # Calculate initial 1-day upfront price for all agents in the fleet
    total_credits = 0
    final_agents = req.agents
    
    machine_credit_costs_per_day = {
        "e2-small": 100,
        "e2-medium": 200,
        "e2-standard-2": 400
    }
    
    for ag in final_agents:
        mtype = ag.machineType if hasattr(ag, 'machineType') and ag.machineType else "e2-small"
        total_credits += machine_credit_costs_per_day.get(mtype, 100)
    
    if req.plan == "starter" and not final_agents:
        raise HTTPException(status_code=400, detail="Must provide at least 1 agent config for starter plan")
    elif req.plan == "medium" and len(final_agents) < 2:
        raise HTTPException(status_code=400, detail="Must provide at least 2 agent configs for medium plan")

    # Pre-provisioning validation: verify all API keys and models before credit deduction or deployment
    from fastapp.models.credentialModel import CredentialModel
    from fastapp.services import validatorService
    import re

    # 1. Back-end Validation: Prevent Duplicates and Circular loops manually sent via API payloads
    role_counts = {}
    for a in final_agents:
        role_upper = a.role.strip().upper()
        if role_upper in role_counts:
            role_counts[role_upper] += 1
            a.role = f"{a.role.strip()} {role_counts[role_upper]}"
        else:
            role_counts[role_upper] = 1

    # Cycle Detection on API payload
    for a in final_agents:
        if not a.reportsTo:
            continue
        visited = set()
        curr = a
        has_cycle = False
        while curr.reportsTo:
            if curr.id in visited:
                has_cycle = True
                break
            visited.add(curr.id)
            parent = next((p for p in final_agents if p.id == curr.reportsTo), None)
            if not parent:
                break
            curr = parent
        if has_cycle:
            a.reportsTo = None

    supported_providers = ["google", "gemini", "openai", "anthropic", "deepseek", "bihand", "custom"]

    for ag in final_agents:
        # Sanitize Skill Names and generate placeholders if custom skills files are empty
        sanitized_files = []
        for file in (ag.skillsFiles or []):
            raw_name = file.get("name") or "unnamed_skill"
            sanitized_name = re.sub(r'[^a-zA-Z0-9_\-]', '_', raw_name).strip().lower()
            if not sanitized_name:
                continue
            content = file.get("content") or ""
            if not content.strip():
                content = f"# {sanitized_name}\nCustom instructions for {sanitized_name}."
            sanitized_files.append({"name": sanitized_name, "content": content})
        ag.skillsFiles = sanitized_files

        provider = ag.provider.lower() if ag.provider else ""
        if provider == "bihand":
            ag.model = "gemini-3.5-flash"
            ag.apiKey = "bihand-system-placeholder"

        if provider not in supported_providers:
            raise HTTPException(
                status_code=400,
                detail=f"Unsupported provider '{ag.provider}' for agent {ag.role}."
            )
            
        if not ag.model or not ag.model.strip():
            raise HTTPException(
                status_code=400,
                detail=f"Model name cannot be empty for agent {ag.role}."
            )

        if provider == "custom" and not (ag.customBaseUrl and ag.customBaseUrl.strip()):
            raise HTTPException(
                status_code=400,
                detail=f"Missing base URL for custom provider agent {ag.role}."
            )

        # A claudecode/codex agent with a subscription auth token (claude setup-token, or a
        # pasted codex ~/.codex/auth.json) bills inference against the user's own plan and
        # never touches the metered-API-key path at all - skip the credential lookup and key
        # validation entirely rather than requiring a throwaway key.
        uses_subscription_auth = ag.agentType in ("claudecode", "codex") and bool(ag.oauthToken)

        if uses_subscription_auth:
            pass
        elif provider == "bihand":
            decrypted_api_key = "bihand-system-placeholder"
        else:
            if not ag.apiKey:
                raise HTTPException(
                    status_code=400,
                    detail=f"Missing API Key credential for agent {ag.role}."
                )

            # Retrieve and decrypt the credential
            creds_doc = CredentialModel.get_by_id(ag.apiKey)
            if not creds_doc:
                raise HTTPException(
                    status_code=400,
                    detail=f"API Key credential with ID '{ag.apiKey}' not found for agent {ag.role}."
                )

            if creds_doc.get("userId") != email:
                raise HTTPException(
                    status_code=403,
                    detail=f"Unauthorized access to credential for agent {ag.role}."
                )

            decrypted_api_key = creds_doc.get("decrypted_data")
            if not decrypted_api_key:
                raise HTTPException(
                    status_code=400,
                    detail=f"Failed to decrypt API Key for agent {ag.role}."
                )

            # validatorService has no vendor endpoint to check a "custom" provider's key
            # against (it's an arbitrary third-party base URL), so skip live validation there
            # - same reasoning as bihand's placeholder key above.
            if provider != "custom":
                is_valid, error_msg = await validatorService.validate_key(provider, decrypted_api_key)
                if not is_valid:
                    raise HTTPException(
                        status_code=400,
                        detail=f"Invalid API Key for agent {ag.role} ({ag.provider}): {error_msg}"
                    )

    # Ensure user has enough credits
    from fastapp.models.userModel import UserModel
    user = UserModel._getUserByEmail(email)
    current_credits = user.get("credits", 0) if user else 0
    if current_credits < total_credits:
        pass  # OSS build: no credit/billing gating (BYOK — bring your own GCP + LLM key)

    # Deduct credits
    tx_details = {
        "action": "provision_company_fleet",
        "fleetName": req.name,
        "plan": req.plan,
        "agentsCount": len(final_agents),
        "agentsLineup": [{
            "role": a.role,
            "title": a.title,
            "machineType": a.machineType,
            "iteration": a.agentType
        } for a in final_agents]
    }
    success = UserModel._deductCredits(email, total_credits, details=tx_details)
    if not success:
         pass  # OSS build: no credit/billing gating (BYOK — bring your own GCP + LLM key)

    # Create Fleet Record
    fleet = FleetModel._create(
        user_id=email,
        name=req.name,
        plan=req.plan,
        total_price=float(total_credits), # Reuse total_price field to store total_credits for compatibility
        agents=[a.dict() for a in final_agents],
        api_budget=float(req.apiBudget),
        mission=req.mission
    )
    
    # Create Initial Task if provided and has valid content
    if req.initialTask and req.initialTask.title.strip():
        from fastapp.models.taskModel import TaskModel
        TaskModel._create(
            fleet_id=fleet["_id"],
            title=req.initialTask.title,
            description=req.initialTask.description,
            status="todo"
        )

    # Kick off Celery task to provision the fleet infrastructure
    from fastapp.tasks import provision_fleet_task
    provision_fleet_task.delay(
        fleet["_id"],
        email,
        req.password
    )
    
    return {
        "message": f"Fleet '{req.name}' is being provisioned.",
        "fleetId": fleet["_id"],
        "dashboardUrl": fleet["bihandUrl"]
    }

@fleetRouter.get("", summary="List user's fleets")
async def list_fleets(auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    fleets = FleetModel._listByUser(email)
    
    from fastapp.database import get_db
    db = get_db()
    
    # Format response
    result = []
    for f in fleets:
        fleet_id = f["_id"]
        instances = list(db["instances"].find({"fleetId": fleet_id, "status": {"$nin": ["deleted"]}}))
        
        infra_burn = 0
        machine_credit_costs_per_day = {
            "e2-micro": 50,
            "e2-small": 100,
            "e2-medium": 200,
            "e2-standard-2": 400
        }
        for inst in instances:
            mtype = inst.get("machineType", "e2-small")
            per_day = machine_credit_costs_per_day.get(mtype, 100)
            infra_burn += per_day
            
        result.append({
            "id": f["_id"],
            "name": f["name"],
            "plan": f["plan"],
            "totalPrice": f["totalPrice"],
            "infraBurn": infra_burn,
            "status": f["status"],
            "dashboardUrl": f["bihandUrl"],
            "createdAt": f["createdAt"],
            "instanceCount": len(instances)
        })
    return result

@fleetRouter.get("/{fleet_id}/export", summary="Export Fleet configuration as a portable JSON template")
async def export_fleet(fleet_id: str, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    fleet = FleetModel._getById(fleet_id)
    
    if not fleet or fleet["userId"] != email:
        raise HTTPException(status_code=404, detail="Fleet not found")
        
    template = FleetModel._exportTemplate(fleet_id)
    return template
@fleetRouter.get("/{fleet_id}", summary="Get fleet details and instances")
async def get_fleet(fleet_id: str, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    
    fleet = FleetModel._getById(fleet_id)
    if not fleet:
        raise HTTPException(status_code=404, detail="Fleet not found")
        
    # Secure Admin backdoor delegation bypass: Admins can view/query fleets of any user
    if fleet["userId"] != email and user_role != "admin":
        raise HTTPException(status_code=404, detail="Fleet not found")
        
    # Get all instances belonging to this fleet
    from fastapp.database import get_db
    collection = get_db()["instances"]
    instances = list(collection.find({"fleetId": fleet_id}))
    
    formatted_instances = []
    for inst in instances:
        formatted_instances.append({
            "id": str(inst["_id"]),
            "role": inst.get("fleetRole", "Unknown"),
            "title": inst.get("title", ""),
            "model": inst.get("model", None),
            "alias": inst.get("alias", ""),
            "status": inst.get("status", "unknown"),
            "ip": inst.get("externalIp", ""),
            "reportsTo": inst.get("reportsTo", None),
            "agentType": inst.get("iteration", ""),
            "avatarHash": inst.get("avatarHash", None),
            "machineType": inst.get("machineType", "e2-small"),
            "token": inst.get("dashboardToken", ""),
            "agentMd": inst.get("agentMd", ""),
            "soulMd": inst.get("soulMd", ""),
            "heartbeatMd": inst.get("heartbeatMd", ""),
            "toolsMd": inst.get("toolsMd", ""),
            "mcpConfig": inst.get("mcpConfig", ""),
            "enabledSkills": inst.get("enabledSkills", []),
            "toolConnections": _sanitized_tool_connections(inst),
            "adapterCapabilities": adapter_capabilities(inst),
            "expiresAt": inst.get("expiresAt"),
            "apiCreditsUsed": inst.get("apiCreditsUsed", 0.0),
        })
        
    infra_burn = 0
    machine_credit_costs_per_day = {
        "e2-micro": 50,
        "e2-small": 100,
        "e2-medium": 200,
        "e2-standard-2": 400
    }
    for inst in instances:
        mtype = inst.get("machineType", "e2-small")
        per_day = machine_credit_costs_per_day.get(mtype, 100)
        infra_burn += per_day

    return {
        "id": str(fleet["_id"]),
        "name": fleet["name"],
        "userId": fleet["userId"],  # Explicitly include the actual owner email for Admin backdoor selector mapping
        "mission": fleet.get("mission", "Execute tasks autonomously."),
        "plan": fleet["plan"],
        "status": fleet["status"],
        "totalPrice": fleet["totalPrice"],
        "infraBurn": infra_burn,
        "dashboardUrl": fleet["bihandUrl"],
        "apiBudget": fleet.get("apiBudget", 0),
        "apiSpend": fleet.get("apiSpend", 0),
        "instances": formatted_instances,
        "expiresAt": instances[0].get("expiresAt") if instances else None
    }

@fleetRouter.post("/{fleet_id}/stop", summary="Stop all agents in fleet")
async def stop_fleet(fleet_id: str, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    fleet = FleetModel._getById(fleet_id)
    if not fleet or (fleet["userId"] != email and user_role != "admin"):
        raise HTTPException(status_code=404, detail="Fleet not found")
        
    if fleet.get("status") in ["deleting", "deleted"]:
        raise HTTPException(status_code=400, detail="Cannot stop a fleet that is being destroyed")
        
    from fastapp.database import get_db
    from fastapp.tasks import stop_instance_task
    instances = list(get_db()["instances"].find({"fleetId": fleet_id}))
    for inst in instances:
        if inst.get("status") == "running":
            InstanceModel._updateStatus(str(inst["_id"]), "stopping_queued")
            stop_instance_task.delay(str(inst["_id"]))
            
    FleetModel._updateStatus(fleet_id, "stopping")
    return {"message": "Fleet stop initiated"}

@fleetRouter.post("/{fleet_id}/start", summary="Start all agents in fleet")
async def start_fleet(fleet_id: str, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    fleet = FleetModel._getById(fleet_id)
    if not fleet or (fleet["userId"] != email and user_role != "admin"):
        raise HTTPException(status_code=404, detail="Fleet not found")
        
    if fleet.get("status") in ["deleting", "deleted"]:
        raise HTTPException(status_code=400, detail="Cannot stop a fleet that is being destroyed")
        
    from fastapp.database import get_db
    from fastapp.tasks import start_instance_task
    instances = list(get_db()["instances"].find({"fleetId": fleet_id}))
    
    from datetime import datetime, timezone
    now = datetime.now(timezone.utc)

    for inst in instances:
        if inst.get("status") == "stopped":
            expires_at = inst.get("expiresAt")
            if expires_at:
                if expires_at.tzinfo is None:
                    expires_at = expires_at.replace(tzinfo=timezone.utc)
                if expires_at < now:
                    continue # Skip expired instances
                    
            InstanceModel._updateStatus(str(inst["_id"]), "provisioning_queued")
            start_instance_task.delay(str(inst["_id"]))
            
    FleetModel._updateStatus(fleet_id, "running")
    return {"message": "Fleet start initiated"}

class ExtendFleetRequest(BaseModel):
    durationMonths: int = Field(..., description="Duration in months to extend (1, 3, 12)")

@fleetRouter.post("/{fleet_id}/extend", summary="Extend fleet duration")
async def extend_fleet(fleet_id: str, req: ExtendFleetRequest, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    fleet = FleetModel._getById(fleet_id)
    if not fleet or (fleet["userId"] != email and user_role != "admin"):
        raise HTTPException(status_code=404, detail="Fleet not found")
        
    from fastapp.models.userModel import UserModel
    from fastapp.database import get_db
    from bson import ObjectId
    from datetime import datetime, timedelta, timezone
    
    machine_credit_costs = {
        "e2-micro": 15,
        "e2-small": 30,
        "e2-medium": 60,
        "e2-standard-2": 120
    }
    
    total_credits_per_month = 0
    instances = list(get_db()["instances"].find({"fleetId": fleet_id}))
    for inst in instances:
        total_credits_per_month += machine_credit_costs.get(inst.get("machineType", "e2-small"), 30)
        
    total_cost = total_credits_per_month * req.durationMonths
    
    owner_email = fleet.get("userId", email)
    user = UserModel._getUserByEmail(owner_email)
    if not user or user.get("credits", 0) < total_cost:
        pass  # OSS build: no credit/billing gating (BYOK — bring your own GCP + LLM key)
        
    if not UserModel._deductCredits(owner_email, total_cost):
        raise HTTPException(status_code=400, detail="Failed to deduct credits.")
        
    new_expires = None
    now = datetime.now(timezone.utc)
    
    for inst in instances:
        current_expires = inst.get("expiresAt")
        if current_expires and current_expires.tzinfo is None:
            current_expires = current_expires.replace(tzinfo=timezone.utc)
            
        base_time = current_expires if current_expires and current_expires > now else now
        new_expires = base_time + timedelta(days=req.durationMonths * 30)
        
        get_db()["instances"].update_one(
            {"_id": ObjectId(inst["_id"])},
            {"$set": {"expiresAt": new_expires, "updatedDate": now}}
        )
        
    return {"message": f"Extended fleet by {req.durationMonths} months.", "newExpiresAt": new_expires}

@fleetRouter.post("/{fleet_id}/instances/{instance_id}/extend", summary="Extend single instance duration")
async def extend_instance(fleet_id: str, instance_id: str, req: ExtendFleetRequest, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    fleet = FleetModel._getById(fleet_id)
    if not fleet or (fleet["userId"] != email and user_role != "admin"):
        raise HTTPException(status_code=404, detail="Fleet not found")
        
    from fastapp.models.userModel import UserModel
    from fastapp.models.instanceModel import InstanceModel
    from fastapp.database import get_db
    from bson import ObjectId
    from datetime import datetime, timedelta, timezone
    
    inst = InstanceModel._getById(instance_id)
    if not inst or inst.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")
        
    machine_credit_costs = {
        "e2-micro": 1500,
        "e2-small": 3000,
        "e2-medium": 6000,
        "e2-standard-2": 12000
    }
    
    monthly_cost = machine_credit_costs.get(inst.get("machineType", "e2-small"), 3000)
    total_cost = monthly_cost * req.durationMonths
    
    owner_email = fleet.get("userId", email)
    user = UserModel._getUserByEmail(owner_email)
    if not user or user.get("credits", 0) < total_cost:
        pass  # OSS build: no credit/billing gating (BYOK — bring your own GCP + LLM key)
        
    if not UserModel._deductCredits(owner_email, total_cost):
        raise HTTPException(status_code=400, detail="Failed to deduct credits.")
        
    now = datetime.now(timezone.utc)
    current_expires = inst.get("expiresAt")
    if current_expires and current_expires.tzinfo is None:
        current_expires = current_expires.replace(tzinfo=timezone.utc)
        
    base_time = current_expires if current_expires and current_expires > now else now
    new_expires = base_time + timedelta(days=req.durationMonths * 30)
    
    get_db()["instances"].update_one(
        {"_id": ObjectId(inst["_id"])},
        {"$set": {"expiresAt": new_expires, "updatedDate": now}}
    )
    
    return {"message": f"Extended instance by {req.durationMonths} months.", "newExpiresAt": new_expires}

class ConnectToolRequest(BaseModel):
    credentialId: str = Field(..., description="The ID of the credential to use")

@fleetRouter.post("/{fleet_id}/instances/{instance_id}/tools/google-workspace/connect", summary="Connect Google Workspace tool using a credential")
async def connect_google_workspace_tool(fleet_id: str, instance_id: str, req: ConnectToolRequest, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    fleet = FleetModel._getById(fleet_id)

    if not fleet or (fleet["userId"] != email and user_role != "admin"):
        raise HTTPException(status_code=404, detail="Fleet not found")

    instance = InstanceModel._getById(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")

    if instance.get("status") != "running":
        raise HTTPException(status_code=400, detail="Forbidden: Agent must be in 'running' status to perform updates.")

    InstanceModel._updateStatus(instance_id, "updating")

    from fastapp.models.credentialModel import CredentialModel
    import json
    
    cred = CredentialModel.get_by_id(req.credentialId)
    if not cred or (cred["userId"] != email and cred["userId"] != fleet["userId"] and user_role != "admin") or cred["type"] != "google_workspace":
        raise HTTPException(status_code=400, detail="Invalid Google Workspace credential")

    try:
        cred_data = json.loads(cred.get("decrypted_data", "{}"))
    except:
        raise HTTPException(status_code=500, detail="Corrupted credential data")

    InstanceModel._setToolConnection(instance_id, "googleWorkspace", {
        "status": "connected",
        "connectedAt": datetime.now(timezone.utc),
        "oauthState": None,
        "email": cred_data.get("email"),
        "scopes": DEFAULT_GOOGLE_SCOPES,
        "credential": {
            "accessToken": cred_data.get("accessToken"),
            "refreshToken": cred_data.get("refreshToken"),
            "tokenType": "Bearer",
            "expiresIn": cred_data.get("expiresIn"),
        },
    })

    # Centralized skill addition to instances
    from fastapp.services.agentProfileService import sync_skills
    enabled_skills = instance.get("enabledSkills", []) or []
    if "bihand-google-workspace" not in enabled_skills:
        enabled_skills.append("bihand-google-workspace")
    
    # We call sync_skills to update the adapter config securely in DB
    next_adapter_config, desired_skills, _ = sync_skills(instance, enabled_skills)
    InstanceModel._setAdapterConfig(instance_id, next_adapter_config)
    
    enabled_skill_slugs = [
        skill.split("/")[-1]
        for skill in desired_skills
        if isinstance(skill, str) and skill.strip() and skill.split("/")[-1] != "bihand"
    ]
    enabled_skill_slugs = list(dict.fromkeys(enabled_skill_slugs))
    InstanceModel._setEnabledSkills(instance_id, enabled_skill_slugs)

    from fastapp.tasks import push_agent_config_task
    push_agent_config_task.delay(
        instance_id,
        instance.get("agentMd", ""),
        instance.get("soulMd", ""),
        instance.get("toolsMd", ""),
        instance.get("mcpConfig", ""),
        enabled_skill_slugs,
    )

    return {"message": "Google Workspace connected successfully"}

@fleetRouter.post("/{fleet_id}/instances/{instance_id}/start", summary="Start specific agent instance")
async def start_fleet_instance(fleet_id: str, instance_id: str, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    fleet = FleetModel._getById(fleet_id)
    if not fleet or (fleet["userId"] != email and user_role != "admin"):
        raise HTTPException(status_code=404, detail="Fleet not found")
        
    instance = InstanceModel._getById(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")
        
    status_val = instance.get("status", "")
    if status_val in ["provisioning_queued", "provisioning", "installing", "starting_queued", "stopping_queued", "restarting_queued", "deleting_queued", "deleting"]:
        raise HTTPException(status_code=400, detail=f"Cannot start agent while status is changing ({status_val})")
        
    if status_val == "running":
        raise HTTPException(status_code=400, detail="Agent is already running")
        
    if status_val not in ["stopped", "error"]:
        raise HTTPException(status_code=400, detail=f"Agent must be stopped or in error to start it (current status: {status_val})")
        
    from datetime import datetime, timezone
    from fastapp.controllers.instanceController import MACHINE_COST_MULTIPLIER
    from fastapp.models.userModel import UserModel
    now = datetime.now(timezone.utc)
    
    last_billed = instance.get("lastBilledAt")
    if last_billed and last_billed.tzinfo is None:
        last_billed = last_billed.replace(tzinfo=timezone.utc)
        
    # Check if the last paid 24h window has already expired
    is_expired = True
    if last_billed:
        duration_since_bill = (now - last_billed).total_seconds()
        if duration_since_bill < 86400: # Within 24-hour paid window
            is_expired = False
            
    if is_expired:
        # Require upfront payment of 1 day to start
        owner_email = fleet.get("userId", email)
        user = UserModel._getUserByEmail(owner_email)
        mtype = instance.get("machineType", "e2-small")
        cost_multiplier = MACHINE_COST_MULTIPLIER.get(mtype, 100)
        
        if not user or user.get("credits", 0) < cost_multiplier:
            pass  # OSS build: no credit/billing gating (BYOK — bring your own GCP + LLM key)
            
        # Deduct credits
        tx_details = {
            "action": "start_agent_instance",
            "instanceId": instance_id,
            "fleetId": fleet_id,
            "vmName": instance.get("vmName"),
            "role": instance.get("fleetRole"),
            "machineType": mtype,
            "iteration": instance.get("iteration")
        }
        if not UserModel._deductCredits(owner_email, cost_multiplier, details=tx_details):
            raise HTTPException(status_code=400, detail="Failed to deduct credits. Try again later.")
            
        # Record new billing timestamp and billingCycleStart
        from bson import ObjectId
        from fastapp.database import get_db
        get_db()["instances"].update_one(
            {"_id": ObjectId(instance_id)},
            {"$set": {"lastBilledAt": now, "billingCycleStart": now, "updatedDate": now}}
        )
            
    from fastapp.tasks import start_instance_task
    try:
        start_instance_task.delay(instance_id, fallback_status="stopped")
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Failed to queue start task, try again: {e}")
    InstanceModel._updateStatus(instance_id, "starting_queued")
    return {"message": "Agent start task queued."}

@fleetRouter.post("/{fleet_id}/instances/{instance_id}/stop", summary="Stop specific agent instance")
async def stop_fleet_instance(fleet_id: str, instance_id: str, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    fleet = FleetModel._getById(fleet_id)
    if not fleet or (fleet["userId"] != email and user_role != "admin"):
        raise HTTPException(status_code=404, detail="Fleet not found")
        
    instance = InstanceModel._getById(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")
        
    status_val = instance.get("status", "")
    if status_val in ["provisioning_queued", "provisioning", "installing", "starting_queued", "stopping_queued", "restarting_queued", "deleting_queued", "deleting"]:
        raise HTTPException(status_code=400, detail=f"Cannot stop agent while status is changing ({status_val})")
        
    if status_val == "stopped":
        raise HTTPException(status_code=400, detail="Agent is already stopped")
        
    if status_val not in ["running", "error"]:
        raise HTTPException(status_code=400, detail=f"Agent must be running or in error to stop it (current status: {status_val})")
        
    from fastapp.tasks import stop_instance_task
    try:
        stop_instance_task.delay(instance_id)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Failed to queue stop task, try again: {e}")
    InstanceModel._updateStatus(instance_id, "stopping_queued")
    return {"message": "Agent stop task queued."}

@fleetRouter.post("/{fleet_id}/instances/{instance_id}/restart", summary="Restart specific agent instance daemon")
async def restart_fleet_instance(fleet_id: str, instance_id: str, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    fleet = FleetModel._getById(fleet_id)
    if not fleet or (fleet["userId"] != email and user_role != "admin"):
        raise HTTPException(status_code=404, detail="Fleet not found")
        
    instance = InstanceModel._getById(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")
        
    status_val = instance.get("status", "")
    if status_val not in ["running", "error"]:
        raise HTTPException(status_code=400, detail="Agent must be running or in error to restart daemon")
        
    from fastapp.tasks import start_instance_task
    try:
        start_instance_task.delay(instance_id, fallback_status="error")
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Failed to queue restart task, try again: {e}")
    InstanceModel._updateStatus(instance_id, "restarting_queued")
    return {"message": "Agent restart task queued."}

@fleetRouter.delete("/{fleet_id}/instances/{instance_id}", summary="Destroy specific agent instance")
async def destroy_fleet_instance(fleet_id: str, instance_id: str, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
    fleet = FleetModel._getById(fleet_id)
    if not fleet or (fleet["userId"] != email and user_role != "admin"):
        raise HTTPException(status_code=404, detail="Fleet not found")
        
    instance = InstanceModel._getById(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")
        
    status_val = instance.get("status", "")
    if status_val in ["provisioning_queued", "provisioning", "installing", "starting_queued", "stopping_queued", "restarting_queued", "deleting_queued", "deleting"]:
        raise HTTPException(status_code=400, detail=f"Cannot destroy agent while status is changing ({status_val})")
        
    from fastapp.tasks import delete_instance_task
    try:
        delete_instance_task.delay(instance_id)
    except Exception as e:
        raise HTTPException(status_code=503, detail=f"Failed to queue destroy task, try again: {e}")
    InstanceModel._updateStatus(instance_id, "deleting_queued")
    return {"message": "Agent destroy task queued."}

class UpgradeInstanceRequest(BaseModel):
    machineType: str = Field(..., description="New GCP Machine Type (e.g., e2-small, e2-medium, e2-standard-2)")
    iteration: str = Field(..., description="New runtime agent type (e.g., opencode, codex, claudecode)")
    provider: str = Field(..., description="New LLM model provider (e.g., anthropic, openai, bihand)")
    model: str = Field(..., description="New LLM model name")
    apiKeyCredentialId: Optional[str] = Field(default=None, description="Optional custom user API key or specific saved credential ID for the new provider")
    oauthToken: Optional[str] = Field(default=None, description="claudecode only: a `claude setup-token` OAuth token to bill inference against the user's own Claude subscription instead of the provider API key. When set, the VM omits ANTHROPIC_API_KEY entirely (it outranks the subscription token in Claude Code's auth precedence).")

@fleetRouter.post("/{fleet_id}/instances/{instance_id}/reconfigure", summary="Reconfigure/Resize Agent VM & Runtime")
async def reconfigure_agent_instance(fleet_id: str, instance_id: str, req: UpgradeInstanceRequest, auth_payload: dict = Depends(get_current_user)):
    email = auth_payload["email"]
    user_role = auth_payload.get("role", "user")
    from fastapp.utils.adminAuth import ADMIN_EMAILS
    if auth_payload.get("email") in ADMIN_EMAILS:
        user_role = "admin"
        
    fleet = FleetModel._getById(fleet_id)
    if not fleet or (fleet["userId"] != email and user_role != "admin"):
        raise HTTPException(status_code=404, detail="Fleet not found")
        
    instance = InstanceModel._getByIdWithKeys(instance_id)
    if not instance or instance.get("fleetId") != fleet_id:
        raise HTTPException(status_code=404, detail="Instance not found")
        
    status_val = instance.get("status", "")
    if status_val in ["provisioning_queued", "provisioning", "installing", "starting_queued", "stopping_queued", "restarting_queued", "deleting_queued", "deleting"]:
        raise HTTPException(status_code=400, detail=f"Cannot reconfigure agent while status is transitioning ({status_val})")

    from fastapp.controllers.instanceController import MACHINE_COST_MULTIPLIER, get_disk_size_gb
    if req.machineType not in MACHINE_COST_MULTIPLIER:
        raise HTTPException(status_code=400, detail="Invalid machine type.")

    # Check and deduct the correct upfront daily credit fee based on the newly requested machineType
    from fastapp.models.userModel import UserModel
    owner_email = fleet.get("userId", email)
    user = UserModel._getUserByEmail(owner_email)
    cost_multiplier = MACHINE_COST_MULTIPLIER[req.machineType]
    
    if not user or user.get("credits", 0) < cost_multiplier:
        pass  # OSS build: no credit/billing gating (BYOK — bring your own GCP + LLM key)
        
    tx_details = {
        "action": "reconfigure_agent_instance",
        "instanceId": instance_id,
        "fleetId": fleet_id,
        "vmName": instance.get("vmName"),
        "machineType": req.machineType,
        "iteration": req.iteration,
        "provider": req.provider
    }
    if not UserModel._deductCredits(owner_email, cost_multiplier, details=tx_details):
        raise HTTPException(status_code=400, detail="Failed to deduct reconfiguration credits. Try again later.")

    # Find the corresponding API Key or Credential Document for the selected provider
    from fastapp.database import get_db
    credential_id = None

    if req.provider == "bihand":
        credential_id = "bihand-system-placeholder"
    elif req.apiKeyCredentialId:
        credential_id = req.apiKeyCredentialId
    else:
        # Fallback: scan user's saved credentials for the requested provider type
        credentials = list(get_db()["credentials"].find({"userId": fleet["userId"], "type": req.provider}))
        if credentials:
            credential_id = str(credentials[0]["_id"])
        else:
            # Re-use the existing instance's API key if it's the exact same provider, otherwise raise error
            if instance.get("provider") == req.provider and instance.get("apiKey"):
                credential_id = instance.get("apiKey")
            else:
                raise HTTPException(
                    status_code=400,
                    detail=f"No saved {req.provider} API credentials found for user. Please save a credential or provide a custom key/credential ID."
                )

    # 1. Update Database parameters immediately, keeping a full audit history of previous VM names for garbage collection
    from bson import ObjectId
    old_vm_name = instance.get("vmName")
    old_zone = instance.get("zone")
    old_machine_type = instance.get("machineType")
    
    reconfig_history_entry = {
        "vmName": old_vm_name,
        "zone": old_zone,
        "machineType": old_machine_type,
        "reconfiguredAt": datetime.now(timezone.utc)
    }

    # Generate a brand new VM name for the resized instance to ensure zero naming collision on GCP
    suffix = secrets.token_hex(4)
    new_vm_name = f"bh-{suffix}" if req.iteration in ["opencode", "claudecode", "codex", "bihand_worker"] else f"nc-{suffix}"
    new_disk_name = f"bh-disk-{suffix}" if req.iteration in ["opencode", "claudecode", "codex", "bihand_worker"] else f"nc-disk-{suffix}"

    get_db()["instances"].update_one(
        {"_id": ObjectId(instance_id)},
        {
            "$set": {
                "vmName": new_vm_name,
                "diskName": new_disk_name,
                "diskSizeGb": get_disk_size_gb(req.machineType),
                "machineType": req.machineType,
                "iteration": req.iteration,
                "provider": req.provider,
                "model": req.model,
                "apiKey": credential_id,
                "oauthToken": req.oauthToken or None,
                "updatedDate": datetime.now(timezone.utc)
            },
            "$push": {
                "vmHistory": reconfig_history_entry
            }
        }
    )

    # 2. Trigger hot-reconfig lifecycle sequence
    from fastapp.tasks import provision_instance_task, reconfigure_gc_and_migrate_workspace_task

    # Instantly offload complete workspace sync and GCP garbage collection to Celery background workers
    reconfigure_gc_and_migrate_workspace_task.delay(
        old_vm_name=old_vm_name,
        old_zone=old_zone,
        instance_id=instance_id,
        old_iteration=instance.get("iteration", "openclaw"),
        new_iteration=req.iteration
    )

    # Fast hot-patch swap: Instantly trigger a rebuild (provision task) to mount the new strategy, machineType, and configuration.
    InstanceModel._updateStatus(instance_id, "provisioning_queued")
    
    provision_instance_task.delay(
        instance_id=instance_id,
        user_id=fleet["userId"],
        provider=req.provider,
        credential_id=str(credential_id),
        password="minerclaw_agent",
        iteration=req.iteration
    )

    return {"message": "Agent reconfiguration and resizing queued successfully."}

